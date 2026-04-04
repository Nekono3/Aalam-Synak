"""
ZipGrade CSV Parser Utility

Parses ZipGrade export files (CSV format) and extracts:
- Student information
- Answers and scores
- Question-level data
"""
import csv
import json
import io
from decimal import Decimal
from typing import List, Dict, Any, Tuple, Optional

from schools.utils import normalize_student_id

import os
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FONT_PATH = os.path.join(BASE_DIR, 'fonts', 'Roboto-Regular.ttf')
FONT_NAME = 'RobotoFlexible'

try:
    pdfmetrics.registerFont(TTFont(FONT_NAME, FONT_PATH))
except Exception:
    pass

class ZipGradeParser:
    """Parser for ZipGrade CSV and XLSX export files."""
    
    # Known column mappings based on ZipGrade export format
    STUDENT_ID_COLUMNS = ['ExternalId', 'External ID', 'ZipGrade ID', 'Student ID', 'StudentId']
    FIRST_NAME_COLUMNS = ['FirstName', 'First Name', 'First']
    LAST_NAME_COLUMNS = ['LastName', 'Last Name', 'Last']
    EARNED_COLUMNS = ['EarnedPts', 'Earned Points', 'Earned', 'Points Earned', 'Score']
    MAX_COLUMNS = ['PossiblePts', 'Possible Points', 'Max Points', 'Possible', 'Max']
    PERCENT_COLUMNS = ['Percent', 'Percentage', 'Pct', '%']
    CLASS_COLUMNS = ['Class', 'Section', 'Period', 'Grade']
    
    def __init__(self, file_content: bytes, encoding: str = 'utf-8', filename: str = ''):
        """Initialize parser with file content.
        
        Args:
            file_content: Raw bytes from uploaded file
            encoding: Character encoding (default utf-8)
            filename: Original filename to detect format
        """
        self.file_content = file_content
        self.encoding = encoding
        self.filename = filename.lower()
        self.headers = []
        self.data = []
        self.answer_columns = []
        self.column_map = {}
    
    def _is_xlsx(self) -> bool:
        """Check if file is an XLSX file."""
        return self.filename.endswith('.xlsx') or self.filename.endswith('.xls')
    
    def _parse_xlsx(self) -> Tuple[List[str], List[Dict[str, str]]]:
        """Parse XLSX file and return headers and rows as dicts."""
        from openpyxl import load_workbook
        
        wb = load_workbook(filename=io.BytesIO(self.file_content), read_only=True, data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        
        # First row is headers
        headers = [str(cell) if cell is not None else '' for cell in rows[0]]
        
        # Convert remaining rows to dicts
        data = []
        for row in rows[1:]:
            row_dict = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = str(cell) if cell is not None else ''
            data.append(row_dict)
        
        wb.close()
        return headers, data
        
    def parse(self) -> Dict[str, Any]:
        """Parse the ZipGrade file and return structured data.
        
        Returns:
            Dictionary containing:
            - total_questions: Number of questions
            - total_students: Number of students
            - results: List of student results
            - errors: List of parsing errors
        """
        errors = []
        results = []
        
        try:
            # Check if XLSX format
            if self._is_xlsx():
                self.headers, rows_data = self._parse_xlsx()
                if not self.headers:
                    return {
                        'total_questions': 0,
                        'total_students': 0,
                        'results': [],
                        'errors': ['Empty or invalid XLSX file.']
                    }
                
                # Map columns and find answer columns
                self._map_columns()
                self._find_answer_columns()
                
                # Extract answer key from PriKey columns (ZipGrade format)
                answer_key = {}
                if hasattr(self, 'prikey_columns') and self.prikey_columns and rows_data:
                    # PriKey values are same for all rows, so just use first row
                    first_row = rows_data[0]
                    for i, col in enumerate(self.prikey_columns, start=1):
                        answer = str(first_row.get(col, '')).strip().upper()
                        if answer:
                            answer_key[str(i)] = answer
                
                # Parse each row
                for row_num, row in enumerate(rows_data, start=2):
                    try:
                        # Skip Answer Key row if present (legacy format)
                        if self._is_answer_key_row(row):
                            if not answer_key:  # Only extract if not already got from PriKey
                                answer_key = self._extract_answer_key(row)
                            continue
                        
                        result = self._parse_row(row)
                        if result:
                            results.append(result)
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                
                
                # Check for max points to fix possible over-detection of columns
                # If we detected way more columns than max points, trust max points
                derived_total_questions = len(self.answer_columns)  # Initialize with current count
                max_values = [r['max_points'] for r in results if r['max_points'] > 0]
                if max_values:
                    from collections import Counter
                    most_common = Counter(max_values).most_common(1)
                    if most_common:
                        likely_max_points = int(most_common[0][0])
                        
                        # If we found significantly more columns than expected, truncate
                        if len(self.answer_columns) > likely_max_points:
                             # Also update the derived count if needed
                            derived_total_questions = likely_max_points
                            # Truncate answer columns to match likely question count
                            self.answer_columns = self.answer_columns[:likely_max_points]
                            # Also truncate answer key
                            answer_key = {k: v for k, v in answer_key.items() if int(k) <= likely_max_points}

                return {
                    'total_questions': derived_total_questions if derived_total_questions > 0 else len(self.answer_columns),
                    'total_students': len(results),
                    'results': results,
                    'errors': errors,
                    'answer_columns': self.answer_columns,
                    'answer_key': answer_key,
                }
            
            # CSV parsing (original logic)
            # Decode file content
            text_content = self.file_content.decode(self.encoding)
            
            # Try to fix common encoding issues
            if '\ufeff' in text_content:
                text_content = text_content.replace('\ufeff', '')
                
        except UnicodeDecodeError:
            # Try different encodings
            for enc in ['utf-8-sig', 'latin-1', 'cp1251', 'cp1252']:
                try:
                    text_content = self.file_content.decode(enc)
                    self.encoding = enc
                    break
                except UnicodeDecodeError:
                    continue
            else:
                return {
                    'total_questions': 0,
                    'total_students': 0,
                    'results': [],
                    'errors': ['Could not decode file. Please ensure it is a valid CSV or XLSX file.']
                }
        
        # Parse CSV
        reader = csv.DictReader(io.StringIO(text_content))
        self.headers = reader.fieldnames or []
        
        if not self.headers:
            return {
                'total_questions': 0,
                'total_students': 0,
                'results': [],
                'errors': ['Empty or invalid CSV file.']
            }
        
        # Map columns
        self._map_columns()
        
        # Find answer columns (usually named like Q1, Q2, etc. or 1, 2, 3)
        self._find_answer_columns()
        
        # Parse each row
        for row_num, row in enumerate(reader, start=2):
            try:
                result = self._parse_row(row)
                if result:
                    results.append(result)
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        
        # Calculate derived total questions from max points if no columns found
        derived_total_questions = len(self.answer_columns)
        if derived_total_questions == 0 and results:
            # Try to infer from max points (assuming 1 point per question)
            max_values = [r['max_points'] for r in results if r['max_points'] > 0]
            if max_values:
                # Use the most common max score or simply the max
                from collections import Counter
                most_common = Counter(max_values).most_common(1)
                if most_common:
                    derived_total_questions = int(most_common[0][0])
        
        return {
            'total_questions': derived_total_questions,
            'total_students': len(results),
            'results': results,
            'errors': errors,
            'answer_columns': self.answer_columns,
        }
    
    def _map_columns(self):
        """Map column names to standardized names."""
        for header in self.headers:
            header_lower = header.lower().strip()
            
            # Student ID
            for col in self.STUDENT_ID_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['student_id'] = header
                    break
            
            # First name
            for col in self.FIRST_NAME_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['first_name'] = header
                    break
            
            # Last name
            for col in self.LAST_NAME_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['last_name'] = header
                    break
            
            # Earned points
            for col in self.EARNED_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['earned'] = header
                    break
            
            # Max points
            for col in self.MAX_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['max'] = header
                    break
            
            # Percentage
            for col in self.PERCENT_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['percent'] = header
                    break
            
            # Class
            for col in self.CLASS_COLUMNS:
                if col.lower() == header_lower:
                    self.column_map['class'] = header
                    break
    
    
    def _find_answer_columns(self):
        """Find columns that contain answer data.
        
        ZipGrade format uses:
        - Stu1, Stu2, ... for student answers
        - PriKey1, PriKey2, ... for correct answers (Answer Key)
        """
        import re
        
        # First, check for ZipGrade format (StuN and PriKeyN columns)
        stu_columns = []
        prikey_columns = []
        
        for header in self.headers:
            if not header:
                continue
            header_str = str(header).strip()
            
            # Match Stu1, Stu2, ... (student answers)
            if re.match(r'^Stu\d+$', header_str):
                stu_columns.append(header_str)
            
            # Match PriKey1, PriKey2, ... (correct answers / answer key)
            elif re.match(r'^PriKey\d+$', header_str):
                prikey_columns.append(header_str)
        
        # Sort by number
        def get_num(col):
            match = re.search(r'(\d+)$', col)
            return int(match.group(1)) if match else 999999
        
        stu_columns.sort(key=get_num)
        prikey_columns.sort(key=get_num)
        
        # If we found ZipGrade format, use it
        if stu_columns:
            self.answer_columns = stu_columns
            self.prikey_columns = prikey_columns
            return
        
        # Fallback to old detection for other formats
        self.prikey_columns = []
        mapped_headers = set(self.column_map.values())
        candidates = []
        
        for header in self.headers:
            if header in mapped_headers:
                continue
            header_stripped = str(header).strip()
            upper_header = header_stripped.upper()
            
            if upper_header in ['DATE', 'TIME', 'SCHOOL', 'CLASS', 'SECTION', 'TEACHER', 'SUBJECT', 'EXAM']:
                continue
            
            # Q1, Q2, etc.
            if re.match(r'^Q\s*[-_]?\s*\d+', upper_header):
                candidates.append(header)
                continue
            
            # Just a digit
            if header_stripped.isdigit():
                candidates.append(header)
                continue
            
            # Ends with digit
            if re.search(r'(\d+)$', header_stripped):
                candidates.append(header)
                continue
        
        candidates.sort(key=get_num)
        self.answer_columns = candidates
    
    def _is_answer_key_row(self, row: Dict[str, str]) -> bool:
        """Check if this row contains the Answer Key.
        
        ZipGrade exports have a row where the name/ID contains 'KEY' or 'ANSWER'.
        This row contains the correct answers for each question.
        """
        # Check student ID field
        if 'student_id' in self.column_map:
            student_id = str(row.get(self.column_map['student_id'], '')).strip().upper()
            if 'KEY' in student_id or 'ANSWER' in student_id or 'ОТВЕТ' in student_id:
                return True
        
        # Check name fields
        if 'first_name' in self.column_map:
            first_name = str(row.get(self.column_map['first_name'], '')).strip().upper()
            if 'KEY' in first_name or 'ANSWER' in first_name or 'ОТВЕТ' in first_name:
                return True
        
        if 'last_name' in self.column_map:
            last_name = str(row.get(self.column_map['last_name'], '')).strip().upper()
            if 'KEY' in last_name or 'ANSWER' in last_name or 'КЛЮЧ' in last_name:
                return True
        
        return False
    
    def _extract_answer_key(self, row: Dict[str, str]) -> Dict[str, str]:
        """Extract answer key from the KEY row.
        
        Returns:
            Dictionary mapping question number to correct answer: {"1": "A", "2": "B", ...}
        """
        answer_key = {}
        for i, col in enumerate(self.answer_columns, start=1):
            answer = str(row.get(col, '')).strip().upper()
            if answer:  # Only include non-empty answers
                answer_key[str(i)] = answer
        return answer_key
    
    def _parse_row(self, row: Dict[str, str]) -> Optional[Dict[str, Any]]:
        """Parse a single row of data.
        
        Args:
            row: Dictionary of column name -> value
            
        Returns:
            Parsed result dictionary or None if row should be skipped
        """
        # Get student ID
        student_id = ''
        if 'student_id' in self.column_map:
            student_id = str(row.get(self.column_map['student_id'], '')).strip()
        
        # Skip empty rows
        if not student_id and not any(row.values()):
            return None
        
        # Use a placeholder for missing student IDs
        if not student_id:
            student_id = 'NO_ID'
        
        # Get names
        first_name = ''
        if 'first_name' in self.column_map:
            first_name = str(row.get(self.column_map['first_name'], '')).strip()
        
        last_name = ''
        if 'last_name' in self.column_map:
            last_name = str(row.get(self.column_map['last_name'], '')).strip()
        
        # Get scores
        earned = Decimal('0')
        if 'earned' in self.column_map:
            try:
                earned_str = str(row.get(self.column_map['earned'], '0')).strip()
                earned = Decimal(earned_str.replace(',', '.')) if earned_str else Decimal('0')
            except:
                pass
        
        max_points = Decimal('0')
        if 'max' in self.column_map:
            try:
                max_str = str(row.get(self.column_map['max'], '0')).strip()
                max_points = Decimal(max_str.replace(',', '.')) if max_str else Decimal('0')
            except:
                pass
        
        # Calculate percentage
        percentage = Decimal('0')
        if 'percent' in self.column_map:
            try:
                pct_str = str(row.get(self.column_map['percent'], '0')).strip()
                pct_str = pct_str.replace('%', '').replace(',', '.')
                percentage = Decimal(pct_str) if pct_str else Decimal('0')
            except:
                pass
        elif max_points > 0:
            percentage = (earned / max_points) * 100
        
        # Get class/section
        class_name = ''
        if 'class' in self.column_map:
            class_name = str(row.get(self.column_map['class'], '')).strip()
        
        # Get answers
        answers = {}
        for i, col in enumerate(self.answer_columns, start=1):
            answer = str(row.get(col, '')).strip().upper()
            answers[str(i)] = answer
        
        return {
            'student_id': student_id,
            'student_id_normalized': normalize_student_id(student_id),
            'first_name': first_name,
            'last_name': last_name,
            'earned': float(earned),
            'max_points': float(max_points),
            'percentage': float(round(percentage, 2)),
            'class_name': class_name,
            'answers': answers,
        }


def calculate_subject_scores(
    answers: Dict[str, str],
    answer_key: Dict[str, str],
    subject_splits: List[Dict],
    student_class_type: str = 'ru',
) -> List[Dict]:
    """Calculate per-subject scores based on question ranges.
    
    Args:
        answers: Student's answers {question_num: answer}
        answer_key: Correct answers {question_num: correct_answer}
        subject_splits: List of {'subject_id': int, 'start': int, 'end': int, 'points': float, 'class_type': str}
        student_class_type: The student's class type ('ru' or 'kg')
        
    Returns:
        List of subject scores
    """
    results = []
    
    for split in subject_splits:
        # Skip splits that don't apply to this student's class type
        split_class_type = split.get('class_type', 'all')
        if split_class_type != 'all' and split_class_type != student_class_type:
            continue
        
        start_q = split['start']
        end_q = split['end']
        points = split.get('points', 1.0)
        
        correct = 0
        total = 0
        question_results = {}
        
        for q_num in range(start_q, end_q + 1):
            q_str = str(q_num)
            student_answer = answers.get(q_str, '')
            correct_answer = answer_key.get(q_str, '')
            
            total += 1
            is_correct = student_answer == correct_answer if correct_answer else False
            question_results[q_str] = {
                'answer': student_answer,
                'correct': correct_answer,
                'is_correct': is_correct
            }
            
            if is_correct:
                correct += 1
        
        earned = correct * points
        max_pts = total * points
        pct = (earned / max_pts * 100) if max_pts > 0 else 0
        
        results.append({
            'subject_id': split['subject_id'],
            'subject_split_id': split.get('split_id'),
            'earned': earned,
            'max_points': max_pts,
            'percentage': round(pct, 2),
            'correct_count': correct,
            'total_count': total,
            'question_results': question_results,
        })
    
    return results


# ==========================================
# Answer Sheet PDF Generator
# ==========================================

def _transliterate(text):
    """Transliterate Cyrillic text to Latin characters.
    Covers Russian and Kyrgyz Cyrillic alphabets.
    """
    CYR_TO_LAT = {
        'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'Yo',
        'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
        'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
        'Ф': 'F', 'Х': 'Kh', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Shch',
        'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya',
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'shch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
        # Kyrgyz-specific characters
        'Ң': 'Ng', 'ң': 'ng',
        'Ү': 'U', 'ү': 'u',
        'Ө': 'O', 'ө': 'o',
    }
    return ''.join(CYR_TO_LAT.get(ch, ch) for ch in str(text))


def _safe_str(value):
    """Convert any value to a clean string. Handles None, NaN, float .0, and 'None'."""
    if value is None:
        return ''
    s = str(value).strip()
    if s.lower() in ('none', 'nan', 'nat', ''):
        return ''
    if s.endswith('.0'):
        s = s[:-2]
    return s


def _clean_student_id(raw_id):
    """Safely clean student ID, preserving leading zeros and handling pandas/openpyxl floats."""
    s = _safe_str(raw_id)
    if not s:
        return '000000000'
    return s.zfill(9)


def _create_overlay(fullname, class_text, student_id):
    """Generate a single transparent PDF page with the student text and ID bubbles.
    Returns a PdfReader object (same pattern as the standalone script).
    """
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfbase import pdfmetrics as pm
    from reportlab.pdfbase.ttfonts import TTFont as TF
    from PyPDF2 import PdfReader

    # Register font every call to be safe (idempotent)
    try:
        pm.registerFont(TF(FONT_NAME, FONT_PATH))
    except Exception:
        pass

    # Calibration constants (match ZipGrade_AnswerSheet.pdf layout)
    START_X = 154.5
    START_Y = 645.5
    COL_SPACING = 15.45
    ROW_SPACING = 17.38
    BUBBLE_RADIUS = 4.2

    packet = io.BytesIO()
    can = rl_canvas.Canvas(packet, pagesize=letter)

    # Use Roboto font (supports Cyrillic, Kyrgyz, English, etc.)
    can.setFont(FONT_NAME, 11)

    can.drawString(149, 696, str(fullname))
    can.drawString(330, 696, str(class_text))

    id_str = _clean_student_id(student_id)

    for col, digit in enumerate(id_str):
        if not digit.isdigit():
            continue
        row = int(digit)
        x = START_X + (col * COL_SPACING)
        y = START_Y - (row * ROW_SPACING)
        can.circle(x, y, BUBBLE_RADIUS, fill=1)

    can.save()
    packet.seek(0)
    return PdfReader(packet)


def generate_answer_sheets(excel_file, template_pdf_path):
    """Generate pre-filled ZipGrade answer sheets from student XLSX data.

    Args:
        excel_file: File-like object or path to an XLSX file with columns:
                    id, name, surname, class, section, class_type
        template_pdf_path: Path to the blank ZipGrade_AnswerSheet.pdf template

    Returns:
        tuple: (BytesIO with the combined PDF, student_count, errors)
    """
    from openpyxl import load_workbook
    from PyPDF2 import PdfReader, PdfWriter

    errors = []

    # Load Excel data
    try:
        if hasattr(excel_file, 'read'):
            wb = load_workbook(filename=io.BytesIO(excel_file.read()), data_only=True)
        else:
            wb = load_workbook(filename=excel_file, data_only=True)
        ws = wb.active
    except Exception as e:
        return None, 0, [f'Failed to read Excel file: {str(e)}']

    # Read headers from first row
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, 0, ['Excel file is empty.']

    raw_headers = rows[0]
    # Normalize headers to lowercase for flexible matching
    headers = []
    for h in raw_headers:
        headers.append(str(h).strip().lower() if h is not None else '')

    # Map expected columns
    col_map = {}
    expected = {
        'id': ['id', 'student_id', 'studentid', 'external_id'],
        'name': ['name', 'first_name', 'firstname'],
        'surname': ['surname', 'last_name', 'lastname'],
        'class': ['class', 'grade'],
        'section': ['section'],
        'class_type': ['class_type', 'classtype', 'type'],
    }

    for field, aliases in expected.items():
        for i, h in enumerate(headers):
            if h in aliases:
                col_map[field] = i
                break

    if 'id' not in col_map:
        return None, 0, ['Could not find "id" column in the Excel file. Expected columns: id, name, surname, class, section, class_type']

    # Process students
    writer = PdfWriter()
    student_count = 0

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            # Extract ALL fields strictly using _safe_str (no field is optional)
            student_id = _safe_str(row[col_map['id']]) if 'id' in col_map and col_map['id'] < len(row) else ''
            name = _safe_str(row[col_map['name']]) if 'name' in col_map and col_map['name'] < len(row) else ''
            surname = _safe_str(row[col_map['surname']]) if 'surname' in col_map and col_map['surname'] < len(row) else ''

            fullname = _transliterate(f"{surname} {name}".strip())

            # Build class text (same strict extraction)
            cls = _safe_str(row[col_map['class']]) if 'class' in col_map and col_map['class'] < len(row) else ''
            section = _safe_str(row[col_map['section']]) if 'section' in col_map and col_map['section'] < len(row) else ''
            class_type = _safe_str(row[col_map['class_type']]) if 'class_type' in col_map and col_map['class_type'] < len(row) else ''

            class_parts = f"{cls} {section}".strip()
            class_text = f"{class_parts} | {class_type}" if class_type and class_parts else (class_type or class_parts)

            # Read template for this student (fresh copy each time)
            reader = PdfReader(template_pdf_path)
            page = reader.pages[0]

            # Create overlay and merge (exactly like standalone script)
            overlay = _create_overlay(fullname, class_text, student_id)
            page.merge_page(overlay.pages[0])
            writer.add_page(page)

            student_count += 1
        except Exception as e:
            errors.append(f'Row {row_idx}: {str(e)}')

    if student_count == 0:
        return None, 0, errors + ['No valid student data found in the file.']

    # Write combined PDF to BytesIO
    output = io.BytesIO()
    writer.write(output)
    output.seek(0)

    wb.close()
    return output, student_count, errors


def generate_answer_sheets_from_school(school, template_pdf_path):
    """Generate pre-filled ZipGrade answer sheets from MasterStudent data.

    STRICT MODE: Never skip a student. If overlay fails, add blank page as fallback.

    Args:
        school: School model instance
        template_pdf_path: Path to the blank ZipGrade_AnswerSheet.pdf template

    Returns:
        tuple: (BytesIO with the combined PDF, student_count, errors)
    """
    from PyPDF2 import PdfReader, PdfWriter
    from schools.models import MasterStudent

    errors = []
    students = list(MasterStudent.objects.filter(school=school).order_by('grade', 'section', 'surname', 'name'))
    total_expected = len(students)

    print(f"[ZipGrade] School: {school.name} | Total students fetched from DB: {total_expected}")

    if total_expected == 0:
        return None, 0, ['No students found for this school.']

    writer = PdfWriter()
    student_count = 0
    failed_count = 0

    for idx, student in enumerate(students, start=1):
        # Extract ALL fields strictly
        name = _safe_str(student.name)
        surname = _safe_str(student.surname)
        grade = _safe_str(student.grade)
        section = _safe_str(student.section)
        sid = _safe_str(student.student_id)

        fullname = _transliterate(f"{surname} {name}".strip())
        class_text = f"{grade} {section}".strip()

        try:
            # Read template (fresh copy each time)
            reader = PdfReader(template_pdf_path)
            page = reader.pages[0]

            # Create overlay and merge
            overlay = _create_overlay(fullname, class_text, sid)
            page.merge_page(overlay.pages[0])
            writer.add_page(page)

            student_count += 1
        except Exception as e:
            # FALLBACK: Still add a page even if overlay fails
            failed_count += 1
            error_msg = f"Student #{idx} {surname} {name} (ID: {sid}): {str(e)}"
            errors.append(error_msg)
            print(f"[ZipGrade] ERROR: {error_msg}")

            try:
                # Add blank template page so student is NOT skipped
                fallback_reader = PdfReader(template_pdf_path)
                fallback_page = fallback_reader.pages[0]

                # Try to at least add ID bubbles
                try:
                    fallback_overlay = _create_overlay(fullname or "ERROR", class_text, sid)
                    fallback_page.merge_page(fallback_overlay.pages[0])
                except Exception:
                    pass  # Even if this fails, we still add the blank page

                writer.add_page(fallback_page)
                student_count += 1
                print(f"[ZipGrade] FALLBACK: Added blank page for student #{idx} {surname} {name}")
            except Exception as fallback_error:
                print(f"[ZipGrade] CRITICAL: Could not add even blank page for student #{idx}: {fallback_error}")
                errors.append(f"CRITICAL - Student #{idx} completely lost: {str(fallback_error)}")

    # Final count validation
    print(f"[ZipGrade] Generation complete: expected={total_expected}, generated={student_count}, failed={failed_count}")
    if student_count != total_expected:
        mismatch_msg = f"MISMATCH: Expected {total_expected} students, generated {student_count} sheets ({total_expected - student_count} missing)"
        print(f"[ZipGrade] WARNING: {mismatch_msg}")
        errors.append(mismatch_msg)

    if student_count == 0:
        return None, 0, errors + ['No student sheets could be generated.']

    output = io.BytesIO()
    writer.write(output)
    output.seek(0)

    return output, student_count, errors
