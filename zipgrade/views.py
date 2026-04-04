import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.db import transaction
from django.core.paginator import Paginator
from django.http import JsonResponse

from accounts.decorators import teacher_or_admin_required, super_admin_required
from schools.models import MasterStudent
from schools.utils import normalize_student_id
from .models import ZipGradeExam, SubjectSplit, ExamResult, SubjectResult
from .forms import ZipGradeUploadForm, SubjectSplitForm
from .utils import ZipGradeParser


@login_required
@teacher_or_admin_required
def upload_view(request):
    """ZipGrade file upload view."""
    if request.method == 'POST':
        form = ZipGradeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            school = form.cleaned_data['school']
            title = form.cleaned_data['title']
            exam_date = form.cleaned_data['exam_date']
            file = form.cleaned_data['file']
            
            # Parse the file
            file_content = file.read()
            parser = ZipGradeParser(file_content, filename=file.name)
            parse_result = parser.parse()
            
            if parse_result['errors'] and not parse_result['results']:
                messages.error(request, _('Failed to parse file: %(errors)s') % {
                    'errors': ', '.join(parse_result['errors'][:3])
                })
                return render(request, 'zipgrade/upload.html', {'form': form})
            
            if not parse_result['results']:
                messages.error(request, _('No student data found in the file.'))
                return render(request, 'zipgrade/upload.html', {'form': form})
            
            # Store in session for preview
            request.session['zipgrade_preview'] = {
                'school_id': school.id,
                'title': title,
                'exam_date': str(exam_date),
                'filename': file.name,
                'parse_result': parse_result,
            }
            request.session.modified = True  # Force session save
            
            return redirect('zipgrade:preview')
    else:
        form = ZipGradeUploadForm()
    
    return render(request, 'zipgrade/upload.html', {'form': form})


@login_required
@teacher_or_admin_required
def preview_view(request):
    """Preview parsed ZipGrade data before saving."""
    preview_data = request.session.get('zipgrade_preview')
    
    if not preview_data:
        messages.warning(request, _('No upload data found. Please upload a file first.'))
        return redirect('zipgrade:upload')
    
    from schools.models import School, Subject
    school = get_object_or_404(School, pk=preview_data['school_id'])
    parse_result = preview_data['parse_result']
    
    # Match students to master list
    matched_results = []
    unknown_count = 0
    
    for result in parse_result['results']:
        normalized_id = result['student_id_normalized']
        
        # Try to find student in master list
        master_student = MasterStudent.objects.filter(
            school=school,
            student_id_normalized=normalized_id
        ).first()
        
        result['matched_student'] = master_student
        result['is_unknown'] = master_student is None
        
        if master_student is None:
            unknown_count += 1
        
        matched_results.append(result)
    
    # Get subjects for subject split dropdown
    subjects = Subject.objects.filter(is_active=True).values('id', 'name')
    subjects_json = json.dumps(list(subjects))
    
    context = {
        'school': school,
        'title': preview_data['title'],
        'exam_date': preview_data['exam_date'],
        'filename': preview_data['filename'],
        'results': matched_results[:50],  # Show first 50 for preview
        'total_students': len(matched_results),
        'total_questions': parse_result['total_questions'],
        'unknown_count': unknown_count,
        'has_more': len(matched_results) > 50,
        'subjects_json': subjects_json,
    }
    
    response = render(request, 'zipgrade/preview.html', context)
    # Disable browser caching for preview page
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@login_required
@teacher_or_admin_required
def confirm_upload_view(request):
    """Confirm and save the ZipGrade data."""
    if request.method != 'POST':
        return redirect('zipgrade:preview')
    
    preview_data = request.session.get('zipgrade_preview')
    if not preview_data:
        messages.warning(request, _('No upload data found. Please upload a file first.'))
        return redirect('zipgrade:upload')
    
    from schools.models import School, Subject
    school = get_object_or_404(School, pk=preview_data['school_id'])
    parse_result = preview_data['parse_result']
    
    # Parse subject splits from POST data
    split_count = int(request.POST.get('split_count', 0))
    print(f"DEBUG: split_count = {split_count}")  # DEBUG
    print(f"DEBUG: POST data = {dict(request.POST)}")  # DEBUG
    subject_splits_data = []
    for i in range(split_count):
        subject_id = request.POST.get(f'split_subject_{i}')
        start_q = request.POST.get(f'split_start_{i}')
        end_q = request.POST.get(f'split_end_{i}')
        print(f"DEBUG: split_{i}: subject={subject_id}, start={start_q}, end={end_q}")  # DEBUG
        if subject_id and start_q and end_q:
            subject_splits_data.append({
                'subject_id': int(subject_id),
                'start_question': int(start_q),
                'end_question': int(end_q),
            })
    print(f"DEBUG: subject_splits_data = {subject_splits_data}")  # DEBUG
    
    try:
        with transaction.atomic():
            # Get extracted answer key from file (if found)
            answer_key = parse_result.get('answer_key', {})
            
            # Create the exam
            exam = ZipGradeExam.objects.create(
                school=school,
                uploaded_by=request.user,
                title=preview_data['title'],
                original_filename=preview_data['filename'],
                exam_date=preview_data['exam_date'],
                total_questions=parse_result['total_questions'],
                total_students=len(parse_result['results']),
                answer_key=json.dumps(answer_key) if answer_key else '',
            )
            
            # Create subject splits
            subject_splits = []
            for split_data in subject_splits_data:
                subject = get_object_or_404(Subject, pk=split_data['subject_id'])
                split = SubjectSplit.objects.create(
                    exam=exam,
                    subject=subject,
                    start_question=split_data['start_question'],
                    end_question=split_data['end_question'],
                )
                subject_splits.append(split)
            
            unknown_count = 0
            
            # Create results
            for result_data in parse_result['results']:
                normalized_id = result_data['student_id_normalized']
                
                # Find matched student
                master_student = MasterStudent.objects.filter(
                    school=school,
                    student_id_normalized=normalized_id
                ).first()
                
                is_unknown = master_student is None
                if is_unknown:
                    unknown_count += 1
                
                # Use update_or_create to handle duplicates in the uploaded file
                exam_result, created = ExamResult.objects.update_or_create(
                    exam=exam,
                    zipgrade_student_id=result_data['student_id'],
                    defaults={
                        'student': master_student,
                        'zipgrade_first_name': result_data['first_name'],
                        'zipgrade_last_name': result_data['last_name'],
                        'earned_points': result_data['earned'],
                        'max_points': result_data['max_points'],
                        'percentage': result_data['percentage'],
                        'answers': json.dumps(result_data['answers']),
                        'is_unknown': is_unknown,
                    }
                )
                
                # Calculate subject results if splits defined
                if subject_splits:
                    answers = result_data['answers']  # Dict: {"1": "A", "2": "B", ...}
                    total_earned = result_data['earned']
                    total_max = result_data['max_points']
                    
                    for split in subject_splits:
                        # Get question range for this subject
                        start_q = split.start_question
                        end_q = split.end_question
                        question_count = split.question_count
                        
                        # Extract answers for this subject's question range
                        subject_answers = {}
                        for q_num in range(start_q, end_q + 1):
                            q_str = str(q_num)
                            if q_str in answers:
                                subject_answers[q_str] = answers[q_str]
                        
                        # Since we don't have answer key in ZipGrade CSV,
                        # do NOT use pro-rated percentages because that creates false data.
                        # Wait for an answer key to be set.
                        earned_points = 0
                        percentage = 0
                        
                        SubjectResult.objects.update_or_create(
                            result=exam_result,
                            subject_split=split,
                            defaults={
                                'earned_points': round(earned_points, 2),
                                'max_points': question_count,
                                'percentage': round(percentage, 2),
                                'question_results': json.dumps(subject_answers),
                            }
                        )
            
            # Update unknown count
            exam.unknown_students = unknown_count
            exam.save()
        
        # Clear session data
        del request.session['zipgrade_preview']
        
        messages.success(request, _('Successfully imported %(count)s student results.') % {
            'count': len(parse_result['results'])
        })
        
        return redirect('zipgrade:exam_detail', pk=exam.pk)
        
    except Exception as e:
        messages.error(request, _('Error saving data: %(error)s') % {'error': str(e)})
        return redirect('zipgrade:preview')


@login_required
@teacher_or_admin_required
def cancel_upload_view(request):
    """Cancel the upload and clear session data."""
    if 'zipgrade_preview' in request.session:
        del request.session['zipgrade_preview']
    return redirect('zipgrade:upload')


from django.db.models import Q

@login_required
@teacher_or_admin_required
def results_view(request):
    """List all ZipGrade exams."""
    exams = ZipGradeExam.objects.all().select_related('school', 'uploaded_by')
    
    # Filter by school for non-super-admins
    if request.user.role != 'super_admin' and request.user.primary_school:
        exams = exams.filter(school=request.user.primary_school)
    
    # Date Filtering
    date_from = request.GET.get('date_from')
    if date_from:
        exams = exams.filter(exam_date__gte=date_from)
        
    date_to = request.GET.get('date_to')
    if date_to:
        exams = exams.filter(exam_date__lte=date_to)
    
    # Search
    search = request.GET.get('search', '')
    if search:
        exams = exams.filter(
            Q(title__icontains=search) |
            Q(original_filename__icontains=search) |
            Q(uploaded_by__first_name__icontains=search) |
            Q(uploaded_by__last_name__icontains=search)
        )
    
    # School filter
    school_filter = request.GET.get('school')
    if school_filter:
        exams = exams.filter(school_id=school_filter)
    
    # Pagination
    paginator = Paginator(exams, 20)
    page = request.GET.get('page', 1)
    exams = paginator.get_page(page)
    
    from schools.models import School
    context = {
        'exams': exams,
        'search': search,
        'school_filter': school_filter,
        'date_from': date_from,
        'date_to': date_to,
        'schools': School.objects.filter(is_active=True),
    }
    
    return render(request, 'zipgrade/results.html', context)


@login_required
@teacher_or_admin_required
def exam_detail_view(request, pk):
    """View exam details and results."""
    exam = get_object_or_404(ZipGradeExam, pk=pk)
    
    results = exam.results.all()
    
    # Filter
    show_unknown = request.GET.get('unknown')
    if show_unknown == '1':
        results = results.filter(is_unknown=True)
    elif show_unknown == '0':
        results = results.filter(is_unknown=False)
    
    # Search
    search = request.GET.get('search', '')
    if search:
        from django.db.models import Q
        results = results.filter(
            Q(zipgrade_first_name__icontains=search) |
            Q(zipgrade_last_name__icontains=search) |
            Q(zipgrade_student_id__icontains=search) |
            Q(student__first_name__icontains=search) |
            Q(student__last_name__icontains=search)
        )
    
    # Sort
    sort = request.GET.get('sort', '-percentage')
    if sort in ['percentage', '-percentage', 'zipgrade_student_id', 'zipgrade_last_name']:
        results = results.order_by(sort)
    
    # Pagination
    paginator = Paginator(results, 50)
    page = request.GET.get('page', 1)
    results = paginator.get_page(page)
    
    context = {
        'exam': exam,
        'results': results,
        'search': search,
        'show_unknown': show_unknown,
        'is_matched_only': show_unknown == '0',
        'is_unknown_only': show_unknown == '1',
        'sort': sort,
        'subject_splits': exam.subject_splits.all(),
    }
    
    return render(request, 'zipgrade/exam_detail.html', context)


@login_required
@teacher_or_admin_required
def add_subject_split_view(request, exam_pk):
    """Add subject split to an exam."""
    exam = get_object_or_404(ZipGradeExam, pk=exam_pk)
    
    if request.method == 'POST':
        form = SubjectSplitForm(request.POST, exam=exam)
        if form.is_valid():
            split = form.save(commit=False)
            split.exam = exam
            split.save()
            
            # Recalculate subject results
            _recalculate_subject_results(exam)
            
            messages.success(request, _('Subject split added successfully.'))
            return redirect('zipgrade:exam_detail', pk=exam.pk)
    else:
        form = SubjectSplitForm(exam=exam)
    
    context = {
        'form': form,
        'exam': exam,
        'title': _('Add Subject Split'),
    }
    
    return render(request, 'zipgrade/subject_split_form.html', context)


@login_required
@teacher_or_admin_required
def edit_subject_split_view(request, pk):
    """Edit a subject split."""
    split = get_object_or_404(SubjectSplit, pk=pk)
    exam = split.exam
    
    if request.method == 'POST':
        form = SubjectSplitForm(request.POST, instance=split, exam=exam)
        if form.is_valid():
            form.save()
            
            # Recalculate subject results
            _recalculate_subject_results(exam)
            
            messages.success(request, _('Subject split updated successfully.'))
            return redirect('zipgrade:exam_detail', pk=exam.pk)
    else:
        form = SubjectSplitForm(instance=split, exam=exam)
    
    context = {
        'form': form,
        'exam': exam,
        'split': split,
        'title': _('Edit Subject Split'),
    }
    
    return render(request, 'zipgrade/subject_split_form.html', context)


@login_required
@teacher_or_admin_required
def delete_subject_split_view(request, pk):
    """Delete a subject split."""
    split = get_object_or_404(SubjectSplit, pk=pk)
    exam = split.exam
    
    if request.method == 'POST':
        # Delete associated subject results
        SubjectResult.objects.filter(subject_split=split).delete()
        split.delete()
        messages.success(request, _('Subject split deleted.'))
        return redirect('zipgrade:exam_detail', pk=exam.pk)
    
    context = {
        'split': split,
        'exam': exam,
    }
    
    return render(request, 'zipgrade/subject_split_confirm_delete.html', context)


def _get_student_class_type(result):
    """Determine a student's class type (ru/kg) from their class info.
    
    Convention:
    - If the student's section contains 'кг' or 'kg' (case-insensitive), they are KG
    - Otherwise, they are RU
    """
    section = ''
    if result.student:
        section = (result.student.section or '').strip().lower()
    elif result.manual_class_name:
        section = result.manual_class_name.strip().lower()
    
    if 'кг' in section or 'kg' in section:
        return 'kg'
    return 'ru'


def _recalculate_subject_results(exam):
    """Recalculate all subject results AND overall ExamResult percentage using answer key."""
    
    splits = list(exam.subject_splits.all())
    
    # Load answer key
    answer_key = {}
    try:
        if exam.answer_key:
            answer_key = json.loads(exam.answer_key)
    except:
        pass
    
    # Determine actual total questions from answer key or exam metadata
    actual_total_questions = len(answer_key) if answer_key else exam.total_questions
    
    for result in exam.results.all():
        try:
            answers = json.loads(result.answers) if result.answers else {}
        except:
            answers = {}
        
        # Determine student's class type for filtering
        student_class_type = _get_student_class_type(result)
        
        # Delete existing subject results for this result
        SubjectResult.objects.filter(result=result).delete()
        
        # Create new subject results for each applicable split
        if splits:
            for split in splits:
                # Skip this split if class type doesn't match
                if split.class_type != 'all' and split.class_type != student_class_type:
                    continue
                
                start_q = split.start_question
                end_q = split.end_question
                question_count = split.question_count
                points_per_q = float(split.points_per_question)
                
                # Count correct answers by comparing against answer key
                correct_count = 0
                question_results = {}
                
                for q_num in range(start_q, end_q + 1):
                    q_str = str(q_num)
                    student_answer = answers.get(q_str, '')
                    correct_answer = answer_key.get(q_str, '')
                    
                    is_correct = (student_answer == correct_answer) if correct_answer else False
                    question_results[q_str] = {
                        'student': student_answer,
                        'correct': correct_answer,
                        'is_correct': is_correct
                    }
                    
                    if is_correct:
                        correct_count += 1
                
                # Calculate earned points based on actual correct answers
                earned = correct_count * points_per_q
                max_pts = question_count * points_per_q
                pct = (correct_count / question_count * 100) if question_count > 0 else 0
                
                SubjectResult.objects.create(
                    result=result,
                    subject_split=split,
                    earned_points=round(earned, 2),
                    max_points=max_pts,
                    percentage=round(pct, 2),
                    question_results=json.dumps(question_results),
                )
        
        # ============================================================
        # CRITICAL: Recalculate overall ExamResult percentage
        # Use answer key to count actual correct answers across ALL questions
        # This replaces ZipGrade's PercentCorrect with real data
        # ============================================================
        if answer_key and actual_total_questions > 0:
            total_correct = 0
            for q_str, correct_answer in answer_key.items():
                student_answer = answers.get(q_str, '')
                if student_answer and student_answer == correct_answer:
                    total_correct += 1
            
            real_percentage = (total_correct / actual_total_questions) * 100
            result.earned_points = total_correct
            result.max_points = actual_total_questions
            result.percentage = round(real_percentage, 2)
            result.save(update_fields=['earned_points', 'max_points', 'percentage'])


@login_required
@teacher_or_admin_required
def delete_exam_view(request, pk):
    """Delete a ZipGrade exam."""
    exam = get_object_or_404(ZipGradeExam, pk=pk)
    
    if request.method == 'POST':
        exam.delete()
        messages.success(request, _('Exam deleted successfully.'))
        return redirect('zipgrade:results')
    
    context = {'exam': exam}
    return render(request, 'zipgrade/exam_confirm_delete.html', context)


@login_required
@teacher_or_admin_required
def edit_unknown_student_view(request, pk):
    """Edit unknown student's manual name or link to existing student."""
    result = get_object_or_404(ExamResult, pk=pk)
    exam = result.exam
    
    if request.method == 'POST':
        manual_first_name = request.POST.get('manual_first_name', '').strip()
        manual_last_name = request.POST.get('manual_last_name', '').strip()
        manual_class_name = request.POST.get('manual_class_name', '').strip()
        link_student_id = request.POST.get('link_student', '').strip()
        
        result.manual_first_name = manual_first_name
        result.manual_last_name = manual_last_name
        result.manual_class_name = manual_class_name
        
        if link_student_id:
            try:
                student = MasterStudent.objects.get(pk=int(link_student_id), school=exam.school)
                result.student = student
                result.is_unknown = False
            except (MasterStudent.DoesNotExist, ValueError):
                pass
        
        result.save()
        
        # Update unknown count if student is now known
        if result.student or result.manual_first_name or result.manual_last_name:
            unknown_count = exam.results.filter(
                is_unknown=True, 
                student__isnull=True, 
                manual_first_name='', 
                manual_last_name=''
            ).count()
            exam.unknown_students = unknown_count
            exam.save()
        
        messages.success(request, _('Student information updated successfully.'))
        return redirect('zipgrade:exam_detail', pk=exam.pk)
    
    # Get available students for linking
    available_students = MasterStudent.objects.filter(school=exam.school).order_by('surname', 'name')
    
    context = {
        'result': result,
        'exam': exam,
        'available_students': available_students,
    }
    
    return render(request, 'zipgrade/edit_unknown_student.html', context)


@login_required
@teacher_or_admin_required
def generate_answersheets_view(request):
    """Upload XLSX and generate pre-filled ZipGrade answer sheets."""
    from schools.models import School

    if request.method == 'POST':
        uploaded_file = request.FILES.get('student_file')

        if not uploaded_file:
            messages.error(request, _('Please select an XLSX file to upload.'))
            return render(request, 'zipgrade/generate_answersheets.html', {'schools': School.objects.filter(is_active=True)})

        # Validate file extension
        if not uploaded_file.name.lower().endswith(('.xlsx', '.xls')):
            messages.error(request, _('Please upload a valid Excel file (.xlsx or .xls).'))
            return render(request, 'zipgrade/generate_answersheets.html', {'schools': School.objects.filter(is_active=True)})

        # Locate template PDF
        import os
        from django.conf import settings
        template_pdf = os.path.join(settings.BASE_DIR, 'static', 'zipgrade', 'ZipGrade_AnswerSheet.pdf')

        if not os.path.exists(template_pdf):
            # Fallback to project root
            template_pdf = os.path.join(settings.BASE_DIR, 'ZipGrade_AnswerSheet.pdf')

        if not os.path.exists(template_pdf):
            messages.error(request, _('ZipGrade answer sheet template not found. Please contact the administrator.'))
            return render(request, 'zipgrade/generate_answersheets.html', {'schools': School.objects.filter(is_active=True)})

        from .utils import generate_answer_sheets

        pdf_output, student_count, errors = generate_answer_sheets(uploaded_file, template_pdf)

        if pdf_output is None:
            error_msg = '; '.join(errors[:5]) if errors else _('Unknown error occurred.')
            messages.error(request, _('Failed to generate answer sheets: %(errors)s') % {'errors': error_msg})
            return render(request, 'zipgrade/generate_answersheets.html', {'schools': School.objects.filter(is_active=True)})

        if errors:
            messages.warning(request, _('Generated %(count)s sheets with %(warn)s warnings: %(errors)s') % {
                'count': student_count,
                'warn': len(errors),
                'errors': '; '.join(errors[:3]),
            })

        # Return PDF as download
        from django.http import HttpResponse
        import datetime

        response = HttpResponse(pdf_output.read(), content_type='application/pdf')
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'ZipGrade_AnswerSheets_{student_count}_students_{timestamp}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    return render(request, 'zipgrade/generate_answersheets.html', {'schools': School.objects.filter(is_active=True)})


@login_required
@teacher_or_admin_required
def generate_answersheets_from_school_view(request):
    """Generate pre-filled ZipGrade answer sheets from MasterStudent data for a school."""
    if request.method != 'POST':
        return redirect('zipgrade:generate_answersheets')

    from schools.models import School
    import os
    import datetime
    from django.conf import settings
    from django.http import HttpResponse
    from .utils import generate_answer_sheets_from_school

    school_id = request.POST.get('school_id')
    if not school_id:
        messages.error(request, _('Please select a school.'))
        return redirect('zipgrade:generate_answersheets')

    school = get_object_or_404(School, pk=school_id, is_active=True)

    # Locate template PDF
    template_pdf = os.path.join(settings.BASE_DIR, 'static', 'zipgrade', 'ZipGrade_AnswerSheet.pdf')
    if not os.path.exists(template_pdf):
        template_pdf = os.path.join(settings.BASE_DIR, 'ZipGrade_AnswerSheet.pdf')
    if not os.path.exists(template_pdf):
        messages.error(request, _('ZipGrade answer sheet template not found. Please contact the administrator.'))
        return redirect('zipgrade:generate_answersheets')

    pdf_output, student_count, errors = generate_answer_sheets_from_school(school, template_pdf)

    if pdf_output is None:
        error_msg = '; '.join(errors[:5]) if errors else _('Unknown error occurred.')
        messages.error(request, _('Failed to generate answer sheets: %(errors)s') % {'errors': error_msg})
        return redirect('zipgrade:generate_answersheets')

    if errors:
        messages.warning(request, _('Generated %(count)s sheets with %(warn)s warnings: %(errors)s') % {
            'count': student_count,
            'warn': len(errors),
            'errors': '; '.join(errors[:3]),
        })

    # Clean school name for filename
    safe_school_name = school.name.replace(' ', '_').replace('/', '-')
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'ZipGrade_{safe_school_name}_{student_count}students_{timestamp}.pdf'

    response = HttpResponse(pdf_output.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@teacher_or_admin_required
def set_answer_key_view(request, pk):
    """Set or update the answer key for an exam."""
    exam = get_object_or_404(ZipGradeExam, pk=pk)
    
    if request.method == 'POST':
        answer_key_input = request.POST.get('answer_key', '').strip()
        
        if answer_key_input:
            # Parse input: support both comma-separated and JSON formats
            try:
                # Try JSON first
                if answer_key_input.startswith('{'):
                    answer_key_dict = json.loads(answer_key_input)
                else:
                    # Comma-separated format: A,B,C,D,...
                    answers_list = [a.strip().upper() for a in answer_key_input.split(',')]
                    answer_key_dict = {str(i+1): ans for i, ans in enumerate(answers_list) if ans}
                
                exam.answer_key = json.dumps(answer_key_dict)
                exam.save()
                
                # Recalculate subject results with actual answer comparison
                _recalculate_subject_results(exam)
                
                messages.success(request, _('Answer key saved. Subject results recalculated for %(count)s students.') % {
                    'count': exam.results.count()
                })
            except json.JSONDecodeError:
                messages.error(request, _('Invalid format. Use comma-separated answers (A,B,C,D,...) or JSON.'))
        else:
            # Clear answer key
            exam.answer_key = ''
            exam.save()
            messages.info(request, _('Answer key cleared.'))
        
        return redirect('zipgrade:exam_detail', pk=exam.pk)
    
    # GET: Display current answer key
    current_key = {}
    try:
        if exam.answer_key:
            current_key = json.loads(exam.answer_key)
    except:
        pass
    
    # Format as comma-separated for display
    if current_key:
        max_q = max(int(k) for k in current_key.keys())
        display_key = ','.join(current_key.get(str(i), '') for i in range(1, max_q + 1))
    else:
        display_key = ''
    
    context = {
        'exam': exam,
        'current_key': display_key,
        'total_questions': exam.total_questions,
    }
    return render(request, 'zipgrade/answer_key_form.html', context)


@login_required
@teacher_or_admin_required
def export_exam_results_excel(request, pk):
    """Export exam results by subject to Excel file.
    
    Format:
    FullName | StudentID | Grade | Section | School | Subject1_TRUE | Subject1_FALSE | Subject2_TRUE | ...
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.utils import timezone
    
    exam = get_object_or_404(ZipGradeExam, pk=pk)
    
    # Recalculate subject results to ensure fresh data
    _recalculate_subject_results(exam)
    
    results = exam.results.all().select_related('student').order_by('student__grade', 'student__section', 'student__surname')
    subject_splits = list(exam.subject_splits.all().select_related('subject').order_by('start_question'))
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Results"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    true_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green for TRUE
    false_header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Red for FALSE
    true_cell_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    false_cell_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Build headers
    base_headers = ['FullName', 'StudentID', 'Grade', 'Section', 'School', 'Total%']
    headers = base_headers.copy()
    
    # Add subject columns: TRUE, FALSE, and % for each subject
    for split in subject_splits:
        headers.append(f"{split.subject.name}_TRUE")
        headers.append(f"{split.subject.name}_FALSE")
        headers.append(f"{split.subject.name}_%")
    
    # Write headers
    percent_header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange for %
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        if col <= len(base_headers):
            cell.fill = header_fill
        elif header.endswith('_TRUE'):
            cell.fill = true_header_fill  # Green for TRUE columns
        elif header.endswith('_FALSE'):
            cell.fill = false_header_fill  # Red for FALSE columns
        elif header.endswith('_%'):
            cell.fill = percent_header_fill  # Orange for % columns
    
    # Write data rows
    row_num = 2
    for result in results:
        # Get student info
        if result.student:
            full_name = result.student.full_name
            student_id = result.student.student_id
            grade = result.student.grade
            section = result.student.section
        else:
            full_name = result.display_name
            student_id = result.zipgrade_student_id
            grade = result.manual_class_name or '-'
            section = '-'
        
        # Base data
        row_data = [
            full_name,
            student_id,
            grade,
            section,
            exam.school.name,
            float(result.percentage)
        ]
        
        # Get subject results
        subject_results = SubjectResult.objects.filter(
            result=result
        ).select_related('subject_split__subject')
        
        # Map by subject split ID
        sr_map = {sr.subject_split_id: sr for sr in subject_results}
        
        # Add subject data
        for split in subject_splits:
            sr = sr_map.get(split.pk)
            if sr:
                # Use stored percentage from SubjectResult
                question_count = split.question_count
                subject_percent = float(sr.percentage)
                
                # Calculate true/false counts based on percentage
                true_count = round(question_count * (subject_percent / 100))
                false_count = question_count - true_count
                
                row_data.append(true_count)
                row_data.append(false_count)
                row_data.append(subject_percent)
            else:
                row_data.append(0)
                row_data.append(split.question_count)
                row_data.append(0.0)
        
        # Write row
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col == 6:  # Total Percentage column
                cell.number_format = '0.0'
            # Color TRUE/FALSE/% columns
            elif col > len(base_headers):
                # Each subject has 3 columns: TRUE, FALSE, %
                relative_pos = (col - len(base_headers) - 1) % 3
                if relative_pos == 0:  # TRUE column
                    cell.fill = true_cell_fill
                elif relative_pos == 1:  # FALSE column
                    cell.fill = false_cell_fill
                else:  # % column
                    cell.number_format = '0.0'
        
        row_num += 1
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)
    
    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"exam_results_{exam.pk}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
