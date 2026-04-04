"""
Utility functions for schools app.
Handles Excel file parsing and Smart ID matching.
"""
import openpyxl
from django.utils.translation import gettext_lazy as _


def parse_master_student_excel(file):
    """
    Parse Master Student List Excel file.
    
    Expected columns (case-insensitive):
    - id / student_id / ID студента
    - name / first_name / имя
    - surname / last_name / фамилия
    - class / grade / класс
    - section / группа
    
    Returns:
        list of dicts with student data
        
    Raises:
        ValueError if file format is invalid
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        
        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).lower().strip())
            else:
                headers.append('')
        
        # Map column names to expected fields
        column_mapping = {}
        
        # Student ID
        for i, h in enumerate(headers):
            if h in ['id', 'student_id', 'studentid', 'id студента', 'ид', 'student id']:
                column_mapping['student_id'] = i
                break
        
        # Name
        for i, h in enumerate(headers):
            if h in ['name', 'first_name', 'firstname', 'имя', 'first name']:
                column_mapping['name'] = i
                break
        
        # Surname
        for i, h in enumerate(headers):
            if h in ['surname', 'last_name', 'lastname', 'фамилия', 'last name']:
                column_mapping['surname'] = i
                break
        
        # Grade/Class
        for i, h in enumerate(headers):
            if h in ['class', 'grade', 'класс', 'class/grade']:
                column_mapping['grade'] = i
                break
        
        # Section
        for i, h in enumerate(headers):
            if h in ['section', 'группа', 'секция', 'group']:
                column_mapping['section'] = i
                break
        
        # Validate required columns
        required = ['student_id', 'name', 'surname', 'grade', 'section']
        missing = [col for col in required if col not in column_mapping]
        
        if missing:
            raise ValueError(
                _('Missing required columns: %(columns)s. Found headers: %(headers)s') % {
                    'columns': ', '.join(missing),
                    'headers': ', '.join(headers)
                }
            )
        
        # Parse rows
        students = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or not any(row):
                continue
            
            try:
                student_id = row[column_mapping['student_id']]
                name = row[column_mapping['name']]
                surname = row[column_mapping['surname']]
                grade = row[column_mapping['grade']]
                section = row[column_mapping['section']]
                
                # Skip if no student ID
                if not student_id:
                    continue
                
                students.append({
                    'student_id': str(student_id).strip(),
                    'name': str(name).strip() if name else '',
                    'surname': str(surname).strip() if surname else '',
                    'grade': str(grade).strip() if grade else '',
                    'section': str(section).strip() if section else '',
                })
            except IndexError:
                continue  # Skip incomplete rows
        
        wb.close()
        return students
        
    except Exception as e:
        if 'Missing required columns' in str(e):
            raise
        raise ValueError(_('Error reading Excel file: %(error)s') % {'error': str(e)})


def normalize_student_id(raw_id):
    """
    Normalize student ID by removing leading zeros.
    Handles both string and integer inputs.
    
    Examples:
        '01251001' -> '1251001'
        '1251001' -> '1251001'
        1251001 -> '1251001'
    """
    try:
        # Convert to string, strip whitespace, convert to int to remove leading zeros
        return str(int(str(raw_id).strip()))
    except (ValueError, TypeError):
        # If conversion fails, just return stripped string
        return str(raw_id).strip()


def find_student_by_id(school, raw_student_id):
    """
    Find a MasterStudent by ID, handling leading zero discrepancies.
    
    Args:
        school: School instance
        raw_student_id: Student ID from ZipGrade (may have different format)
    
    Returns:
        MasterStudent instance or None if not found
    """
    from .models import MasterStudent
    
    normalized_id = normalize_student_id(raw_student_id)
    
    # Try to find by normalized ID
    student = MasterStudent.objects.filter(
        school=school,
        student_id_normalized=normalized_id
    ).first()
    
    return student
