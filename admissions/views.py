import json
from django.shortcuts import render, get_object_or_404, redirect, reverse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse, JsonResponse
from django.db import models, transaction
from django.db.models import Count, Q, Avg
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from .models import (
    AdmissionCycle, ExternalSchool, AdmissionCandidate,
    AdmissionResult, AdmissionSubjectScore, AdmissionUploadSession,
    AdmissionRegistration, AdmissionQuestion, AdmissionQuestionOption,
    AdmissionSubjectSplit, AdmissionMasterAnswer, OnlineAttempt, OnlineAttemptAnswer,
    RoundResultSession, RoundResult, CycleLink, ExamRecording, OnlineExamViolation
)
from .forms import (
    AdmissionXLSXUploadForm, AdmissionRegistrationForm,
    AdmissionQuestionForm, AdmissionQuestionOptionFormSet,
    AdmissionSubjectSplitFormSet
)
from schools.models import School
from exams.models import OnlineExam, ExamAttempt
import openpyxl

def is_super_admin(user):
    return user.is_authenticated and user.role == 'super_admin'

def is_student(user):
    return user.is_authenticated and user.role == 'student'


def _normalize_variant(raw_variant):
    """Normalize variant string: convert Cyrillic А→A, Б→B to ensure consistent matching.
    
    The XLSX files use Cyrillic letters (1А, 1Б, 2А, 2Б) but the DB stores
    Latin letters (1A, 1B, 2A, 2B). They look identical but are different Unicode chars.
    """
    if not raw_variant:
        return ""
    # Map Cyrillic to Latin equivalents
    CYRILLIC_TO_LATIN = {
        'А': 'A', 'а': 'A',  # Cyrillic A
        'Б': 'B', 'б': 'B',  # Cyrillic Be
        'В': 'B', 'в': 'B',  # Cyrillic Ve (also commonly maps to B in variant context)
    }
    result = ""
    for ch in str(raw_variant).strip():
        result += CYRILLIC_TO_LATIN.get(ch, ch)
    return result.upper()


def _auto_create_external_school(school_name):
    """Auto-create an ExternalSchool record if it doesn't exist.
    
    Uses case-insensitive matching to avoid duplicates.
    Auto-generates a school_id like 'AUTO-1', 'AUTO-2', etc.
    """
    if not school_name or not school_name.strip():
        return None
    
    normalized = school_name.strip()
    
    # Case-insensitive lookup
    existing = ExternalSchool.objects.filter(name__iexact=normalized).first()
    if existing:
        return existing
    
    # Generate unique school_id correctly regardless of string sorting (e.g. AUTO-10 vs AUTO-2)
    auto_ids = ExternalSchool.objects.filter(
        school_id__startswith='AUTO-'
    ).values_list('school_id', flat=True)
    
    max_num = 0
    for sid in auto_ids:
        try:
            num = int(sid.replace('AUTO-', ''))
            if num > max_num:
                max_num = num
        except ValueError:
            pass
            
    new_num = max_num + 1
    
    school = ExternalSchool.objects.create(
        name=normalized,
        school_id=f'AUTO-{new_num}'
    )
    return school


@login_required
@user_passes_test(is_super_admin)
def admission_dashboard(request):
    """Admin dashboard for admissions."""
    cycles = AdmissionCycle.objects.all()
    total_candidates = AdmissionCandidate.objects.count()
    context = {
        'cycles': cycles,
        'total_candidates': total_candidates,
    }
    return render(request, 'admissions/dashboard.html', context)

@login_required
@user_passes_test(is_super_admin)
def cycle_list(request):
    """List of all admission cycles."""
    cycles = AdmissionCycle.objects.all()
    return render(request, 'admissions/cycle_list.html', {'cycles': cycles})

@login_required
@user_passes_test(is_super_admin)
def cycle_create(request):
    """Create a new admission cycle."""
    if request.method == 'POST':
        name = request.POST.get('name')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        is_active = request.POST.get('is_active') == 'on'
        prevent_back_navigation = request.POST.get('prevent_back_navigation') == 'on'
        passing_score = request.POST.get('passing_score', 60)
        
        timer_minutes = request.POST.get('timer_minutes', 60)
        require_webcam = request.POST.get('require_webcam') == 'on'
        enable_tab_warnings = request.POST.get('enable_tab_warnings') == 'on'
        max_tab_switches = request.POST.get('max_tab_switches', 3)
        
        cycle = AdmissionCycle.objects.create(
            name=name,
            start_date=start_date,
            end_date=end_date,
            is_active=is_active,
            prevent_back_navigation=prevent_back_navigation,
            passing_score=passing_score,
            timer_minutes=timer_minutes,
            require_webcam=require_webcam,
            enable_tab_warnings=enable_tab_warnings,
            max_tab_switches=max_tab_switches
        )
        
        # Save links
        order = 0
        for i in range(50):
            title = request.POST.get(f'link_title_{i}', '').strip()
            url = request.POST.get(f'link_url_{i}', '').strip()
            target_gender = request.POST.get(f'link_gender_{i}', 'all').strip()
            if title and url:
                CycleLink.objects.create(cycle=cycle, title=title, url=url, target_gender=target_gender, order=order)
                order += 1
            
        messages.success(request, _('Admission cycle created successfully.'))
        return redirect('admissions:cycle_list')
    
    return render(request, 'admissions/cycle_form.html')

@login_required
@user_passes_test(is_super_admin)
def cycle_edit(request, pk):
    """Edit an existing admission cycle."""
    cycle = get_object_or_404(AdmissionCycle, pk=pk)
    if request.method == 'POST':
        cycle.name = request.POST.get('name')
        cycle.start_date = request.POST.get('start_date')
        cycle.end_date = request.POST.get('end_date')
        cycle.is_active = request.POST.get('is_active') == 'on'
        cycle.prevent_back_navigation = request.POST.get('prevent_back_navigation') == 'on'
        cycle.passing_score = request.POST.get('passing_score', 60)
        
        cycle.timer_minutes = request.POST.get('timer_minutes', 60)
        cycle.require_webcam = request.POST.get('require_webcam') == 'on'
        cycle.enable_tab_warnings = request.POST.get('enable_tab_warnings') == 'on'
        cycle.max_tab_switches = request.POST.get('max_tab_switches', 3)
        
        cycle.save()
        
        # Re-create links
        cycle.links.all().delete()
        order = 0
        for i in range(50):
            title = request.POST.get(f'link_title_{i}', '').strip()
            url = request.POST.get(f'link_url_{i}', '').strip()
            target_gender = request.POST.get(f'link_gender_{i}', 'all').strip()
            if title and url:
                CycleLink.objects.create(cycle=cycle, title=title, url=url, target_gender=target_gender, order=order)
                order += 1
        
        messages.success(request, _('Admission cycle updated successfully.'))
        return redirect('admissions:cycle_list')
    
    existing_links = list(cycle.links.all().values('title', 'url', 'target_gender'))
    return render(request, 'admissions/cycle_form.html', {
        'cycle': cycle, 
        'is_edit': True,
        'existing_links_json': json.dumps(existing_links, ensure_ascii=False),
    })

# ============================================================
# CYCLE EXAM CONFIGURATION VIEWS
# ============================================================

@login_required
@user_passes_test(is_super_admin)
def cycle_questions(request, pk):
    """View and manage questions for a cycle's admission exam."""
    cycle = get_object_or_404(AdmissionCycle, pk=pk)
    questions = cycle.admission_questions.prefetch_related('options').all()
    return render(request, 'admissions/cycle_questions.html', {
        'cycle': cycle, 'questions': questions
    })


@login_required
@user_passes_test(is_super_admin)
def cycle_add_question(request, pk):
    """Add a question to a cycle's exam."""
    cycle = get_object_or_404(AdmissionCycle, pk=pk)
    
    if request.method == 'POST':
        form = AdmissionQuestionForm(request.POST, request.FILES)
        formset = AdmissionQuestionOptionFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            question = form.save(commit=False)
            question.cycle = cycle
            if not question.order:
                last_order = cycle.admission_questions.order_by('-order').values_list('order', flat=True).first()
                question.order = (last_order or 0) + 1
            question.save()
            
            formset.instance = question
            formset.save()
            
            messages.success(request, _('Question added successfully.'))
            return redirect('admissions:cycle_questions', pk=cycle.pk)
    else:
        form = AdmissionQuestionForm()
        formset = AdmissionQuestionOptionFormSet()
    
    return render(request, 'admissions/cycle_question_form.html', {
        'cycle': cycle, 'form': form, 'formset': formset
    })


@login_required
@user_passes_test(is_super_admin)
def cycle_edit_question(request, pk, question_pk):
    """Edit an existing question."""
    cycle = get_object_or_404(AdmissionCycle, pk=pk)
    question = get_object_or_404(AdmissionQuestion, pk=question_pk, cycle=cycle)
    
    if request.method == 'POST':
        form = AdmissionQuestionForm(request.POST, request.FILES, instance=question)
        formset = AdmissionQuestionOptionFormSet(request.POST, instance=question)
        
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, _('Question updated successfully.'))
            return redirect('admissions:cycle_questions', pk=cycle.pk)
    else:
        form = AdmissionQuestionForm(instance=question)
        formset = AdmissionQuestionOptionFormSet(instance=question)
    
    return render(request, 'admissions/cycle_question_form.html', {
        'cycle': cycle, 'form': form, 'formset': formset,
        'question': question, 'is_edit': True
    })


@login_required
@user_passes_test(is_super_admin)
def cycle_delete_question(request, pk, question_pk):
    """Delete a question from a cycle."""
    cycle = get_object_or_404(AdmissionCycle, pk=pk)
    question = get_object_or_404(AdmissionQuestion, pk=question_pk, cycle=cycle)
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, _('Question deleted.'))
        return redirect('admissions:cycle_questions', pk=cycle.pk)
    
    return render(request, 'admissions/cycle_question_delete.html', {
        'cycle': cycle, 'question': question
    })


@login_required
@user_passes_test(is_super_admin)
def cycle_copy_questions(request, pk):
    """Copy all questions (with options) from another cycle into this one."""
    cycle = get_object_or_404(AdmissionCycle, pk=pk)
    other_cycles = AdmissionCycle.objects.exclude(pk=pk).annotate(
        q_count=Count('admission_questions')
    ).filter(q_count__gt=0)

    if request.method == 'POST':
        source_id = request.POST.get('source_cycle')
        if source_id:
            source_cycle = get_object_or_404(AdmissionCycle, pk=source_id)
            source_questions = source_cycle.admission_questions.prefetch_related('options').all()

            # Determine starting order
            last_order = cycle.admission_questions.order_by('-order').values_list('order', flat=True).first() or 0

            copied = 0
            with transaction.atomic():
                for sq in source_questions:
                    last_order += 1
                    new_q = AdmissionQuestion.objects.create(
                        cycle=cycle,
                        question_text=sq.question_text,
                        order=last_order,
                        points=sq.points,
                    )
                    # Copy image file if exists
                    if sq.question_image:
                        new_q.question_image.save(
                            sq.question_image.name.split('/')[-1],
                            sq.question_image.file,
                            save=True
                        )
                    for opt in sq.options.all():
                        AdmissionQuestionOption.objects.create(
                            question=new_q,
                            text=opt.text,
                            is_correct=opt.is_correct,
                            order=opt.order,
                        )
                    copied += 1

            messages.success(request, _('Copied %(count)d questions from "%(source)s".') % {
                'count': copied, 'source': source_cycle.name
            })
            return redirect('admissions:cycle_questions', pk=cycle.pk)

    return render(request, 'admissions/cycle_copy_questions.html', {
        'cycle': cycle, 'other_cycles': other_cycles
    })


@login_required
@user_passes_test(is_super_admin)
def cycle_export_questions_xlsx(request, pk):
    """Export all questions for a cycle to an .xlsx file."""
    cycle = get_object_or_404(AdmissionCycle, pk=pk)
    questions = cycle.admission_questions.prefetch_related('options').order_by('order', 'id')

    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    headers = ['Order', 'Question Text', 'Points',
               'Option A', 'Option B', 'Option C', 'Option D',
               'Option E', 'Option F', 'Correct Option']
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    LETTERS = 'ABCDEF'

    for q in questions:
        opts = list(q.options.order_by('order', 'id'))
        row = [q.order, q.question_text, q.points]
        correct_letter = ''
        for i in range(6):
            if i < len(opts):
                row.append(opts[i].text)
                if opts[i].is_correct:
                    correct_letter = LETTERS[i]
            else:
                row.append('')
        row.append(correct_letter)
        ws.append(row)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_name = cycle.name.replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="Questions_{safe_name}.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(is_super_admin)
def cycle_import_questions_xlsx(request, pk):
    """Import questions from an uploaded .xlsx file into a cycle."""
    cycle = get_object_or_404(AdmissionCycle, pk=pk)

    if request.method == 'POST' and request.FILES.get('file'):
        uploaded = request.FILES['file']
        try:
            workbook = openpyxl.load_workbook(uploaded, data_only=True)
            sheet = workbook.active

            last_order = cycle.admission_questions.order_by('-order').values_list('order', flat=True).first() or 0
            imported = 0
            LETTERS = 'ABCDEF'

            with transaction.atomic():
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                    # Expect: Order, Question Text, Points, OptA..OptF, Correct
                    if not row or len(row) < 5:
                        continue
                    question_text = str(row[1] or '').strip()
                    if not question_text:
                        continue

                    order_val = row[0]
                    try:
                        order_val = int(order_val)
                    except (TypeError, ValueError):
                        last_order += 1
                        order_val = last_order

                    points_val = row[2]
                    try:
                        points_val = int(points_val)
                    except (TypeError, ValueError):
                        points_val = 1

                    new_q = AdmissionQuestion.objects.create(
                        cycle=cycle,
                        question_text=question_text,
                        order=order_val,
                        points=points_val,
                    )

                    # Parse options (columns 3..8) and correct letter (column 9)
                    correct_letter = str(row[9] if len(row) > 9 and row[9] else '').strip().upper()

                    for i in range(6):
                        col_idx = 3 + i
                        if col_idx < len(row) and row[col_idx] and str(row[col_idx]).strip():
                            opt_text = str(row[col_idx]).strip()
                            is_correct = (LETTERS[i] == correct_letter)
                            AdmissionQuestionOption.objects.create(
                                question=new_q,
                                text=opt_text,
                                is_correct=is_correct,
                                order=i,
                            )

                    if order_val > last_order:
                        last_order = order_val
                    imported += 1

            messages.success(request, _('Imported %(count)d questions from Excel.') % {'count': imported})
            return redirect('admissions:cycle_questions', pk=cycle.pk)

        except Exception as e:
            messages.error(request, _('Error importing file: %(error)s') % {'error': str(e)})

    return render(request, 'admissions/cycle_import_questions.html', {
        'cycle': cycle,
    })


@login_required
@user_passes_test(is_super_admin)
def cycle_subject_splits(request, pk):
    """Configure subject ranges for a cycle calculation."""
    cycle = get_object_or_404(AdmissionCycle, pk=pk)
    
    # Universal splits are stored under variant '1A'
    queryset = AdmissionSubjectSplit.objects.filter(cycle=cycle, variant='1A')
    
    if request.method == 'POST':
        formset = AdmissionSubjectSplitFormSet(request.POST, instance=cycle, queryset=queryset)
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.variant = '1A'
                instance.save()
            for obj in formset.deleted_objects:
                obj.delete()
            messages.success(request, _('Subject splits saved successfully.'))
            return redirect('admissions:cycle_questions', pk=cycle.pk)
    else:
        formset = AdmissionSubjectSplitFormSet(instance=cycle, queryset=queryset)
    
    return render(request, 'admissions/cycle_subject_splits.html', {
        'cycle': cycle, 'formset': formset
    })



@login_required
@user_passes_test(is_super_admin)
def school_registry(request):
    """List of external schools for the registry (auto-populated from registrations and uploads)."""
    active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
    
    # Get all ExternalSchool records with counts from results and registrations
    schools = ExternalSchool.objects.annotate(
        result_count=Count(
            'id',
            filter=Q(id__in=ExternalSchool.objects.filter(
                name__in=AdmissionResult.objects.filter(cycle=active_cycle).values_list('school_name', flat=True)
            ).values_list('id', flat=True))
        ) if active_cycle else Count('id', filter=Q(pk=None)),
    )
    
    # Build a dict of school_name -> student counts from AdmissionResult
    school_counts = {}
    if active_cycle:
        for row in AdmissionResult.objects.filter(cycle=active_cycle).values('school_name').annotate(cnt=Count('id')):
            school_counts[row['school_name'].strip().lower()] = row['cnt']
    
    # Also count from AdmissionRegistration
    reg_counts = {}
    if active_cycle:
        for row in AdmissionRegistration.objects.filter(cycle=active_cycle).values('school_name').annotate(cnt=Count('id')):
            reg_counts[row['school_name'].strip().lower()] = row['cnt']
    
    # Attach counts to each school
    for school in schools:
        key = school.name.strip().lower()
        school.student_result_count = school_counts.get(key, 0)
        school.student_registration_count = reg_counts.get(key, 0)
        school.total_student_count = school.student_result_count + school.student_registration_count
    
    context = {
        'schools': schools,
        'active_cycle': active_cycle,
    }
    return render(request, 'admissions/school_registry.html', context)

@login_required
@user_passes_test(is_super_admin)
def external_school_create(request):
    """Add a new external school to the registry."""
    if request.method == 'POST':
        name = request.POST.get('name')
        school_id = request.POST.get('school_id')
        
        if name and school_id:
            ExternalSchool.objects.create(name=name, school_id=school_id)
            messages.success(request, _('External school added to registry.'))
            return redirect('admissions:school_registry')
        else:
            messages.error(request, _('Please provide both name and school ID.'))
            
    return render(request, 'admissions/school_registry_form.html')
@login_required
@user_passes_test(is_super_admin)
def external_school_delete(request, pk):
    """Delete an external school from the registry."""
    school = get_object_or_404(ExternalSchool, pk=pk)
    if request.method == 'POST':
        school.delete()
        messages.success(request, _('School removed from registry.'))
    return redirect('admissions:school_registry')

@login_required
@user_passes_test(is_super_admin)
def region_list(request):
    """List all regions and handle adding new ones."""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            from .models import Region
            Region.objects.get_or_create(name=name)
            messages.success(request, _('Region added successfully.'))
        return redirect('admissions:region_list')
        
    from .models import Region
    regions = Region.objects.all()
    return render(request, 'admissions/region_list.html', {'regions': regions})

@login_required
@user_passes_test(is_super_admin)
def region_edit(request, pk):
    """Edit a region's name."""
    from .models import Region
    region = get_object_or_404(Region, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            region.name = name
            region.save()
            messages.success(request, _('Region updated successfully.'))
    return redirect('admissions:region_list')

@login_required
@user_passes_test(is_super_admin)
def region_delete(request, pk):
    """Delete a region."""
    from .models import Region
    region = get_object_or_404(Region, pk=pk)
    if request.method == 'POST':
        region.delete()
        messages.success(request, _('Region deleted successfully.'))
    return redirect('admissions:region_list')

@login_required
@user_passes_test(is_super_admin)
def candidate_list(request):
    """List of all candidates with filtering and exam status."""
    search_query = request.GET.get('search', '')
    cycle_id = request.GET.get('cycle')
    
    candidates = AdmissionCandidate.objects.select_related(
        'user', 'cycle', 'previous_school'
    ).all()
    
    if search_query:
        candidates = candidates.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    if cycle_id:
        from .models import AdmissionRegistration
        registered_user_ids = AdmissionRegistration.objects.filter(cycle_id=cycle_id).values_list('user_id', flat=True)
        attempt_user_ids = OnlineAttempt.objects.filter(cycle_id=cycle_id).values_list('student_id', flat=True)
        candidates = candidates.filter(
            Q(cycle_id=cycle_id) | 
            Q(user_id__in=registered_user_ids) | 
            Q(user_id__in=attempt_user_ids)
        ).distinct()
    
    # Efficiently fetch online attempts and results
    student_ids = [c.user_id for c in candidates]
    cycle_ids = list(set([c.cycle_id for c in candidates if c.cycle_id]))
    candidate_ids = [c.id for c in candidates]
    
    attempts = OnlineAttempt.objects.filter(
        student_id__in=student_ids,
        cycle_id__in=cycle_ids
    )
    attempt_map = {(a.student_id, a.cycle_id): a for a in attempts}
    
    results = AdmissionResult.objects.filter(candidate_id__in=candidate_ids)
    result_map = {r.candidate_id: r for r in results}
    
    for candidate in candidates:
        if candidate.cycle_id:
            attempt = attempt_map.get((candidate.user_id, candidate.cycle_id))
            result = result_map.get(candidate.id)
            
            candidate.online_attempt = attempt
            candidate.admission_result = result
            
            if result and result.percentage is not None:
                candidate.is_passed = result.percentage >= (candidate.cycle.passing_score if candidate.cycle else 0)
    
    all_cycles = AdmissionCycle.objects.all()
    active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
    
    context = {
        'candidates': candidates,
        'all_cycles': all_cycles,
        'selected_cycle_id': cycle_id,
        'search_query': search_query,
        'active_cycle': active_cycle,
    }
    return render(request, 'admissions/candidate_list.html', context)

@login_required
@user_passes_test(is_student)
def student_admission_view(request):
    """View for students to see their admission status and take the admission exam."""
    user = request.user
    candidate = None
    try:
        candidate = user.admission_profile
    except AdmissionCandidate.DoesNotExist:
        pass
        
    # Find all cycles the student has already participated in (active or inactive)
    participated_cycle_ids = set(
        OnlineAttempt.objects.filter(student=user).values_list('cycle_id', flat=True)
    )
    if candidate:
        result_cycle_ids = set(
            AdmissionResult.objects.filter(candidate=candidate).values_list('cycle_id', flat=True)
        )
        participated_cycle_ids.update(result_cycle_ids)
        
    if participated_cycle_ids:
        # If they participated in anything, show ONLY those cycles, regardless of is_active
        active_cycles = list(AdmissionCycle.objects.filter(id__in=participated_cycle_ids).order_by('-start_date', '-created_at'))
    else:
        # Otherwise, show currently active cycles
        active_cycles = list(AdmissionCycle.objects.filter(is_active=True).order_by('-start_date', '-created_at'))
    
    cycles_data = []
    for active_cycle in active_cycles:
        has_online_exam = False
        admission_attempt = None
        admission_result = None
        
        if active_cycle.admission_questions.exists():
            has_online_exam = True
            admission_attempt = OnlineAttempt.objects.filter(
                cycle=active_cycle, student=user
            ).first()
            admission_result = AdmissionResult.objects.filter(
                cycle=active_cycle, candidate__user=user
            ).first()
            
        cycles_data.append({
            'cycle': active_cycle,
            'has_online_exam': has_online_exam,
            'admission_attempt': admission_attempt,
            'admission_result': admission_result,
        })
        
    context = {
        'candidate': candidate,
        'cycles_data': cycles_data,
    }
    return render(request, 'admissions/student_admission.html', context)

@login_required
@user_passes_test(is_super_admin)
def admission_analytics(request):
    """Old analytics view - redirecting to new consolidated dashboard."""
    return redirect('admissions:admission_analytics_dashboard')


@login_required
@user_passes_test(is_super_admin)
def export_admission_results(request, cycle_id=None, admission_type=None):
    """Export admission results for a cycle to XLSX.
    
    If admission_type is provided ('online' or 'offline'), filters by that type
    and auto-names the file accordingly.
    """
    if cycle_id:
        cycle = get_object_or_404(AdmissionCycle, pk=cycle_id)
    else:
        cycle = AdmissionCycle.objects.filter(is_active=True).first()
        if not cycle:
            cycle = AdmissionCycle.objects.first()
            
    if not cycle:
        messages.error(request, _("No admission cycle found to export."))
        return redirect('admissions:admission_analytics_dashboard')

    # Fetch results, optionally filtered by admission type
    results = AdmissionResult.objects.filter(cycle=cycle).order_by('-score')
    if admission_type and admission_type in ('online', 'offline'):
        results = results.filter(admission_type=admission_type)

    subject_names = list(
        AdmissionSubjectScore.objects.filter(
            result__cycle=cycle
        ).values_list('subject_name', flat=True).distinct().order_by('subject_name')
    )

    wb = Workbook()
    
    # -------------------------------------------------------------
    # SHEET 1: Student Results
    # -------------------------------------------------------------
    ws_students = wb.active
    ws_students.title = str(_("All Students"))

    headers_students = [
        str(_("ФИО")),
        str(_("Пол")),
        str(_("Школа")),
        str(_("Предыдущая школа")),
        str(_("Регион / Город")),
        str(_("Адрес")),
        str(_("Телефон 1")),
        str(_("Телефон 2")),
    ]
    
    survey_qs = AdmissionQuestion.objects.filter(cycle=cycle, is_survey_question=True).order_by('order')
    survey_headers = [str(sq.question_text[:30]) for sq in survey_qs]
    headers_students += survey_headers
    
    for subj in subject_names:
        headers_students.append(subj)
    headers_students += [
        str(_("Балл")),
        str(_("Процент")),
        str(_("Медаль")),
        str(_("Источник")),
        str(_("Длительность")),
        str(_("Предупреждения")),
        str(_("Статус экзамена")),
    ]
    ws_students.append(headers_students)

    header_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    for cell in ws_students[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    result_ids = list(results.values_list('id', flat=True))
    all_subject_scores = AdmissionSubjectScore.objects.filter(result_id__in=result_ids)
    score_map = {}
    for ss in all_subject_scores:
        score_map.setdefault(ss.result_id, {})[ss.subject_name] = ss.correct_count

    MEDAL_LABELS = {'gold': str(_("Золото")), 'silver': str(_("Серебро")), 'bronze': str(_("Бронза"))}
    GENDER_LABELS = {'M': str(_("М")), 'F': str(_("Ж"))}

    # Map registrations
    all_regs = list(AdmissionRegistration.objects.filter(cycle=cycle))
    # Collect candidate data manually for missing info
    
    school_stats = {} # For Sheet 2
    
    for result in results:
        gender_display = ''
        prev_school = ''
        address = ''
        city_town = result.region
        
        # Pull detailed data from linked candidate or matched registration
        if result.candidate:
            gender_display = GENDER_LABELS.get(result.candidate.gender, '')
            prev_school = result.candidate.previous_school.name if result.candidate.previous_school else result.school_name
            address = result.candidate.address
        
        # Fallback to result's fields (which should have been populated from registration)
        if not prev_school:
            # Check if we can find it in Registration
            reg = AdmissionRegistration.objects.filter(user=result.candidate.user, cycle=cycle).first() if (result.candidate and result.candidate.user) else None
            if reg:
                if not gender_display: gender_display = GENDER_LABELS.get(reg.gender, '')
                # Note: Registration model doesn't have address or prev_school, 
                # but candidate_detail shows it in result.candidate properties.
        
        # Ensure we have a display for city_town
        city_town = result.region or city_town
        
        # If still empty, try to match by name (legacy/backup)
        if not gender_display or not city_town:
            matched_reg = next((r for r in all_regs if r.full_name.strip().lower() == f"{result.first_name} {result.last_name}".strip().lower()), None)
            if matched_reg:
                if not gender_display: gender_display = GENDER_LABELS.get(matched_reg.gender, '')
                if not city_town: city_town = matched_reg.region
                
        row = [
            f"{result.first_name} {result.last_name}",
            gender_display,
            result.school_name,
            prev_school,
            city_town,
            address,
            result.phone1 or '',
            result.phone2 or '',
        ]
        
        # Add survey answers
        if hasattr(result, 'exam_attempt') and result.exam_attempt:
            attempt_answers = result.exam_attempt.answers.all()
            ans_map = {ans.question_id: ans.selected_option.text if ans.selected_option else "" for ans in attempt_answers}
            for sq in survey_qs:
                row.append(ans_map.get(sq.id, ""))
        else:
            for sq in survey_qs:
                row.append("")
        
        # Sheet 2 Prep: School Stats
        school = result.school_name.strip()
        if school:
            if school not in school_stats:
                school_stats[school] = {'count': 0, 'total_score': 0, 'subject_scores': {s: 0 for s in subject_names}}
            school_stats[school]['count'] += 1
            school_stats[school]['total_score'] += result.percentage
        
        subj_scores = score_map.get(result.id, {})
        for subj in subject_names:
            val = subj_scores.get(subj, 0)
            row.append(val)
            if school:
                school_stats[school]['subject_scores'][subj] += val
                
        row += [
            float(result.score) if result.score else 0.0,
            round(result.percentage, 1) if result.percentage else 0.0,
            MEDAL_LABELS.get(result.medal, '—'),
            result.admission_type.title() if result.admission_type else '',
            result.duration_display,
            result.tab_switch_count,
            result.exam_attempt.get_status_display() if hasattr(result, 'exam_attempt') and result.exam_attempt else "N/A"
        ]
        ws_students.append(row)

    for col in ws_students.columns:
        ws_students.column_dimensions[col[0].column_letter].width = 15

    # -------------------------------------------------------------
    # SHEET 2: School Analytics
    # -------------------------------------------------------------
    ws_schools = wb.create_sheet(title=str(_("School Analytics")))
    headers_schools = [str(_("Школа")), str(_("Кол-во учеников")), str(_("Средний %"))] + subject_names
    ws_schools.append(headers_schools)
    
    for cell in ws_schools[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
    for school_name, stats in sorted(school_stats.items(), key=lambda x: x[1]['total_score']/x[1]['count'], reverse=True):
        count = stats['count']
        avg_pct = round(stats['total_score'] / count, 1)
        row = [school_name, count, avg_pct]
        for subj in subject_names:
            row.append(round(stats['subject_scores'][subj] / count, 1))
        ws_schools.append(row)
        
    for col in ws_schools.columns:
        ws_schools.column_dimensions[col[0].column_letter].width = 18

    # -------------------------------------------------------------
    # SHEET 3: Subject Rankings
    # -------------------------------------------------------------
    ws_subjects = wb.create_sheet(title=str(_("Subject Rankings")))
    headers_subjects = [str(_("Предмет")), str(_("1 место")), str(_("2 место")), str(_("3 место"))]
    ws_subjects.append(headers_subjects)
    
    for cell in ws_subjects[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
    for subj in subject_names:
        # Get top 3
        top_scores = AdmissionSubjectScore.objects.filter(
            subject_name=subj, result__cycle=cycle
        ).order_by('-correct_count', '-result__score')[:3]
        
        row = [subj]
        for ts in top_scores:
            row.append(f"{ts.result.first_name} {ts.result.last_name} ({ts.percentage}%)")
        # Pad if less than 3
        while len(row) < 4:
            row.append("—")
        ws_subjects.append(row)

    for col in ws_subjects.columns:
        ws_subjects.column_dimensions[col[0].column_letter].width = 25

    if admission_type and admission_type in ('online', 'offline'):
        filename = f"{admission_type}.xlsx"
    else:
        filename = f"Results_{cycle.name.replace(' ', '_')}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response


@login_required
@user_passes_test(is_super_admin)
def admission_analytics_upload(request):
    """Upload and process offline admission XLSX data."""
    if request.method == 'POST':
        form = AdmissionXLSXUploadForm(request.POST, request.FILES)
        if form.is_valid():
            cycle = form.cleaned_data['cycle']
            selected_region = form.cleaned_data['region']
            admission_type = form.cleaned_data['admission_type']
            uploaded_file = request.FILES['file']
            filename = uploaded_file.name
            
            final_region = selected_region
            
            # Save session
            session = AdmissionUploadSession.objects.create(
                cycle=cycle,
                region=final_region,
                admission_type=admission_type,
                file=uploaded_file,
                uploaded_by=request.user
            )
            
            # Ensure subject splits exist for this cycle
            _ensure_subject_splits(cycle)
            
            # Process XLSX
            try:
                workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
                sheet = workbook.active
                
                # Dynamic Header Detection
                headers = {}
                header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                
                surname_mapping = ['фамилия', 'last name', 'surname']
                given_mapping = ['имя', 'first name', 'given name']
                variant_mapping = ['вариант', 'variant']
                school_mapping = ['школа', 'school', 'лицей', 'гимназия']
                answers_mapping = ['ответы', 'answers', 'answer']
                phone1_mapping = ['телефон 1', 'phone 1', 'contact 1', 'телефон', 'phone']
                phone2_mapping = ['телефон 2', 'phone 2', 'contact 2']
                score_mapping = ['балл', 'score', 'результат', 'points']
                school_id_mapping = ['id ученика', 'student id', 'id', '№']
                gender_mapping = ['пол', 'gender', 'sex']
                
                for idx, cell in enumerate(header_row):
                    if not cell: continue
                    cell_lower = str(cell).lower().strip()
                    
                    if any(kw in cell_lower for kw in surname_mapping):
                        headers['last_name'] = idx
                    elif any(kw in cell_lower for kw in given_mapping):
                        headers['first_name'] = idx
                    elif any(kw in cell_lower for kw in variant_mapping):
                        headers['variant'] = idx
                    elif any(kw in cell_lower for kw in answers_mapping):
                        headers['answers'] = idx
                    elif any(kw in cell_lower for kw in phone1_mapping) and 'phone1' not in headers:
                        headers['phone1'] = idx
                    elif any(kw in cell_lower for kw in phone2_mapping):
                        headers['phone2'] = idx
                    elif any(kw in cell_lower for kw in school_mapping) and 'school' not in headers:
                        headers['school'] = idx
                    elif any(kw in cell_lower for kw in score_mapping) and 'score' not in headers:
                        headers['score'] = idx
                    elif any(kw in cell_lower for kw in school_id_mapping) and 'student_id' not in headers:
                        headers['student_id'] = idx
                
                # Load master answers for this cycle
                master_answers = {}
                for ma in AdmissionMasterAnswer.objects.filter(cycle=cycle):
                    master_answers[ma.variant] = ma.answers[:70]  # Only first 70
                
                # Load universal subject splits (stored as 1A)
                universal_splits = list(AdmissionSubjectSplit.objects.filter(cycle=cycle, variant='1A'))
                
                rows_processed = 0
                rows_skipped = 0
                
                for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                    try:
                        # Extract name parts
                        first_name = ""
                        last_name = ""
                        
                        if 'last_name' in headers:
                            last_name = str(row[headers['last_name']]).strip() if row[headers['last_name']] else ""
                        if 'first_name' in headers:
                            first_name = str(row[headers['first_name']]).strip() if row[headers['first_name']] else ""
                        
                        if not first_name and not last_name:
                            continue
                        
                        # Extract variant
                        variant = ""
                        if 'variant' in headers and row[headers['variant']]:
                            variant = _normalize_variant(row[headers['variant']])
                        
                        # Extract school
                        school_name = "N/A"
                        if 'school' in headers and headers['school'] < len(row) and row[headers['school']]:
                            school_name = str(row[headers['school']]).strip()
                        
                        # Auto-create ExternalSchool from uploaded data
                        if school_name and school_name != "N/A":
                            _auto_create_external_school(school_name)
                        
                        # Extract answer string
                        answer_string = ""
                        if 'answers' in headers and headers['answers'] < len(row) and row[headers['answers']]:
                            answer_string = str(row[headers['answers']]).strip()
                        
                        # Extract phones
                        phone1 = str(row[headers['phone1']]).strip() if 'phone1' in headers and headers['phone1'] < len(row) and row[headers['phone1']] else None
                        phone2 = str(row[headers['phone2']]).strip() if 'phone2' in headers and headers['phone2'] < len(row) and row[headers['phone2']] else None
                        
                        # Take only first 70 characters of the answer string
                        answers_70 = answer_string[:70] if answer_string else ""
                        
                        # Check against master answer
                        correct_count = 0
                        wrong_count = 0
                        
                        if variant in master_answers and answers_70:
                            master = master_answers[variant]
                            for i in range(min(70, len(answers_70), len(master))):
                                student_char = answers_70[i]
                                master_char = master[i]
                                # Dash and star are treated as wrong/skipped
                                if student_char in ('-', '*', ' '):
                                    wrong_count += 1
                                elif student_char == master_char:
                                    correct_count += 1
                                else:
                                    wrong_count += 1
                            # If answer string is shorter than 70, remaining are wrong
                            if len(answers_70) < 70:
                                wrong_count += 70 - len(answers_70)
                        else:
                            # No master answer available — try to use score column
                            if 'score' in headers and headers['score'] < len(row) and row[headers['score']]:
                                try:
                                    correct_count = int(float(str(row[headers['score']])))
                                    wrong_count = 70 - correct_count
                                except (ValueError, TypeError):
                                    pass
                        
                        score_val = float(correct_count)
                        percentage = (correct_count / 70 * 100) if correct_count > 0 else 0
                        
                        # Create result
                        result = AdmissionResult.objects.create(
                            cycle=cycle,
                            first_name=first_name,
                            last_name=last_name,
                            school_name=school_name,
                            region=final_region,
                            variant=variant,
                            answer_string=answer_string,
                            score=score_val,
                            total_questions=70,
                            percentage=percentage,
                            correct_count=correct_count,
                            wrong_count=wrong_count,
                            phone1=phone1,
                            phone2=phone2,
                            is_passed=percentage >= cycle.passing_score,
                            admission_type=admission_type,
                            upload_session=session
                        )
                        
                        # Calculate subject scores
                        if variant in master_answers:
                            _calculate_subject_scores(result, answers_70, master_answers[variant], universal_splits)
                        
                        rows_processed += 1
                    except Exception as row_err:
                        rows_skipped += 1
                        continue
                
                msg = _('Successfully processed %d records.') % rows_processed
                if rows_skipped:
                    msg += _(' Skipped %d rows.') % rows_skipped
                messages.success(request, msg)
                return redirect('admissions:admission_analytics_dashboard')
                
            except Exception as e:
                messages.error(request, _('Error processing file: %s') % str(e))
                session.delete()
    else:
        form = AdmissionXLSXUploadForm()
        
    return render(request, 'admissions/analytics_upload.html', {'form': form})


def _ensure_subject_splits(cycle):
    """Auto-create subject splits for all 4 variants if none exist."""
    if AdmissionSubjectSplit.objects.filter(cycle=cycle).exists():
        return
    
    # Subject arrangement per spec
    SPLITS = {
        '1A': [
            ('математика', 1, 20),
            ('кыргыз тил', 21, 30),
            ('биология', 31, 40),
            ('география', 41, 50),
            ('тарых', 51, 60),
            ('англис тил', 61, 70),
        ]
    }
    
    for variant, subjects in SPLITS.items():
        for subj_name, start, end in subjects:
            AdmissionSubjectSplit.objects.create(
                cycle=cycle,
                variant=variant,
                subject_name=subj_name,
                start_question=start,
                end_question=end,
            )


def _calculate_subject_scores(result, answers_70, master_answer, splits):
    """Calculate per-subject scores for a student result."""
    from .models import AdmissionSubjectScore
    
    for split in splits:
        correct = 0
        wrong = 0
        start = split.start_question - 1  # Convert to 0-indexed
        end = split.end_question  # exclusive end for slicing
        
        for i in range(start, min(end, 70)):
            if i < len(answers_70) and i < len(master_answer):
                student_char = answers_70[i]
                master_char = master_answer[i]
                if student_char in ('-', '*', ' '):
                    wrong += 1
                elif student_char == master_char:
                    correct += 1
                else:
                    wrong += 1
            else:
                wrong += 1
        
        total_q = split.end_question - split.start_question + 1
        
        AdmissionSubjectScore.objects.create(
            result=result,
            subject_name=split.subject_name,
            correct_count=correct,
            wrong_count=wrong,
            total_questions=total_q,
        )


@login_required
@user_passes_test(is_super_admin)
def delete_admission_session(request, pk):
    """Delete an upload session and all its associated results."""
    session = get_object_or_404(AdmissionUploadSession, pk=pk)
    if request.method == 'POST':
        session.delete()
        messages.success(request, _('Upload session and all associated results deleted successfully.'))
    return redirect('admissions:admission_analytics_dashboard')


@login_required
@user_passes_test(is_super_admin)
def admission_analytics_dashboard(request):
    """Main dashboard for consolidated admission analytics."""
    cycle_id = request.GET.get('cycle')
    if cycle_id:
        active_cycle = get_object_or_404(AdmissionCycle, pk=cycle_id)
    else:
        active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
        if not active_cycle:
            active_cycle = AdmissionCycle.objects.first()
            
    all_cycles = AdmissionCycle.objects.all().order_by('-start_date')
        
    results = AdmissionResult.objects.filter(cycle=active_cycle) if active_cycle else AdmissionResult.objects.all()
    
    # Filters & Configurable Passing Score
    region_filter = request.GET.get('region')
    type_filter = request.GET.get('admission_type')
    
    try:
        custom_passing_score = int(request.GET.get('passing_score', active_cycle.passing_score if active_cycle else 60))
    except (ValueError, TypeError):
        custom_passing_score = active_cycle.passing_score if active_cycle else 60
        
    if region_filter:
        results = results.filter(region=region_filter)
    if type_filter:
        results = results.filter(admission_type=type_filter)
        
    # Totals
    total_offline = AdmissionResult.objects.filter(cycle=active_cycle, admission_type='offline').count() if active_cycle else 0
    total_online = AdmissionResult.objects.filter(cycle=active_cycle, admission_type='online').count() if active_cycle else 0

    # Dashboard calculations
    total_students = results.count()
    passed_students = results.filter(percentage__gte=custom_passing_score).count()
    pass_rate = (passed_students / total_students * 100) if total_students > 0 else 0
    avg_score = results.aggregate(Avg('percentage'))['percentage__avg'] or 0

    medal_stats = {
        'gold': results.filter(percentage__gte=75).count(),
        'silver': results.filter(percentage__gte=60, percentage__lt=75).count(),
        'bronze': results.filter(percentage__gte=50, percentage__lt=60).count(),
    }
    
    top_students = results.order_by('-percentage', '-score')[:10]
    
    # School ranking (offline & online combined)
    top_schools = (results.values('school_name')
                   .annotate(avg_score=Avg('percentage'), count=Count('id'))
                   .filter(count__gte=3)  # Only rank schools with at least 3 participants
                   .order_by('-avg_score')[:10])
                   
    # Region participation
    region_stats = (results.values('region')
                    .annotate(count=Count('id'), avg=Avg('percentage'))
                    .order_by('-count'))
                    
    # Online vs Offline comparison
    mode_comparison = {
        'offline_avg': results.filter(admission_type='offline').aggregate(Avg('percentage'))['percentage__avg'] or 0,
        'online_avg': results.filter(admission_type='online').aggregate(Avg('percentage'))['percentage__avg'] or 0,
    }

    context = {
        'all_cycles': all_cycles,
        'active_cycle': active_cycle,
        'custom_passing_score': custom_passing_score,
        'total_students': total_students,
        'total_online': total_online,
        'total_offline': total_offline,
        'passed_students': passed_students,
        'pass_rate': pass_rate,
        'avg_score': avg_score,
        'medal_stats': medal_stats,
        'top_students': top_students,
        'top_schools': top_schools,
        'region_stats': region_stats,
        'mode_comparison': mode_comparison,
        'results': results[:50], 
        'regions': list(AdmissionResult.objects.filter(cycle=active_cycle).order_by('region').values_list('region', flat=True).distinct()) if active_cycle else [],
        'selected_region': region_filter,
        'selected_type': type_filter,
        'passing_score': custom_passing_score,
        'upload_sessions': AdmissionUploadSession.objects.filter(cycle=active_cycle) if active_cycle else [],
        'has_master_answers': AdmissionMasterAnswer.objects.filter(cycle=active_cycle).exists() if active_cycle else False,
    }
    return render(request, 'admissions/analytics_dashboard.html', context)


@login_required
@user_passes_test(is_super_admin)
def download_sample_xlsx(request):
    """Generate and serve a sample XLSX template for admission uploads."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    headers = ["Фамилия", "Имя", "Вариант", "Пол", "Школа", "№ ID ученика", "Телефон 1", "Телефон 2", "Ответы"]
    ws.append(headers)
    ws.append(["Асанов", "Улук", "1A", "164", "Бала", "9 9", "0505552132", "", "БГВВГВАВГГГАГББВВА..."])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Admission_Template.xlsx"'
    request.session.pop('export_files', None)
    return response


# ============================================================
# STUDENT ONLINE EXAM VIEWS
# ============================================================

from datetime import timedelta
from django.utils import timezone
from .models import OnlineAttempt, OnlineAttemptAnswer, AdmissionQuestionOption

@login_required
@user_passes_test(is_student)
def cycle_exam_start(request, pk):
    """Student starts an online admission exam."""
    cycle = get_object_or_404(AdmissionCycle, pk=pk, is_active=True)
    if not cycle.admission_questions.exists():
        messages.error(request, _('No questions available for this cycle.'))
        return redirect('admissions:student_admission')
        
    # Check if they have already started/finished ANY OTHER active cycle
    active_cycles = AdmissionCycle.objects.filter(is_active=True).exclude(pk=cycle.pk)
    has_other_attempts = OnlineAttempt.objects.filter(student=request.user, cycle__in=active_cycles).exists()
    
    has_other_results = False
    try:
        candidate = request.user.admission_profile
        has_other_results = AdmissionResult.objects.filter(candidate=candidate, cycle__in=active_cycles).exists()
    except AdmissionCandidate.DoesNotExist:
        pass
        
    if has_other_attempts or has_other_results:
        messages.error(request, _('You have already participated in another admission exam. You cannot take multiple exams.'))
        return redirect('admissions:student_admission')
        
    attempt, created = OnlineAttempt.objects.get_or_create(
        cycle=cycle,
        student=request.user,
    )
    
    if attempt.status == 'completed':
        messages.info(request, _('You have already completed this exam.'))
        return redirect('admissions:student_admission')
        
    return redirect('admissions:cycle_exam_take', attempt_pk=attempt.pk)

@login_required
@user_passes_test(is_student)
def cycle_exam_take(request, attempt_pk):
    """Student taking the online admission exam."""
    attempt = get_object_or_404(OnlineAttempt, pk=attempt_pk, student=request.user)
    
    if attempt.status == 'locked':
        messages.error(request, _('Your exam was automatically submitted due to violating proctoring rules.'))
        return redirect('admissions:cycle_exam_submit', attempt_pk=attempt.pk)
        
    if attempt.status != 'in_progress':
        return redirect('admissions:student_admission')
        
    # Check timeout
    if attempt.started_at:
        duration = timedelta(minutes=attempt.cycle.timer_minutes)
        if timezone.now() > attempt.started_at + duration:
            attempt.status = 'timed_out'
            attempt.save()
            messages.warning(request, _('Time is up! Your exam was automatically submitted.'))
            # Just redirect to submit which will handle missing POST data to calculate 0 points for unanswered.
            return redirect('admissions:cycle_exam_submit', attempt_pk=attempt.pk)
    else:
        attempt.started_at = timezone.now()
        attempt.save()

    # Get student's variant from registration
    registration = AdmissionRegistration.objects.filter(user=request.user, cycle=attempt.cycle).first()
    variant = registration.variant if registration else '1A'
    splits = attempt.cycle.subject_splits.filter(variant='1A').order_by('start_question')
    
    # Pre-calculate assigned subject for each question
    question_list = list(attempt.cycle.admission_questions.prefetch_related('options').order_by('order'))
    for i, q in enumerate(question_list):
        q_num = i + 1
        q.subject_name = ""
        q.original_num = q_num
        for split in splits:
            if split.start_question <= q_num <= split.end_question:
                q.subject_name = split.subject_name
                break

    import random
    
    survey_questions = []
    shuffled_question_list = []
    current_subject = None
    current_group = []
    
    for q in question_list:
        opts = list(q.options.all())
        q.shuffled_opts_list = opts
        
        if getattr(q, 'is_survey_question', False):
            survey_questions.append(q)
            continue
            
        random.shuffle(opts)
        if q.subject_name != current_subject:
            if current_group:
                random.shuffle(current_group)
                shuffled_question_list.extend(current_group)
            current_subject = q.subject_name
            current_group = [q]
        else:
            current_group.append(q)
            
    if current_group:
        random.shuffle(current_group)
        shuffled_question_list.extend(current_group)
        
    question_list = survey_questions + shuffled_question_list

    # Get existing answers to restore them
    existing_answers = {a.question_id: a.selected_option_id for a in attempt.answers.all()}
    
    context = {
        'attempt': attempt,
        'cycle': attempt.cycle,
        'questions': question_list,
        'existing_answers': existing_answers,
        'splits': splits,
        # Proctoring settings
        'require_webcam': attempt.cycle.require_webcam,
        'enable_tab_warnings': attempt.cycle.enable_tab_warnings,
        'max_tab_switches': attempt.cycle.max_tab_switches,
    }
    return render(request, 'admissions/cycle_exam_take.html', context)

@login_required
@require_POST
def save_violation_ajax(request, attempt_pk):
    """AJAX: Record a proctoring violation (tab switch)."""
    attempt = get_object_or_404(OnlineAttempt, pk=attempt_pk, student=request.user)
    
    if attempt.status != 'in_progress':
        return JsonResponse({'status': 'error', 'message': 'Exam not in progress'}, status=400)

    try:
        data = json.loads(request.body)
        violation_type = data.get('type', 'tab_switch')
    except json.JSONDecodeError:
        violation_type = 'tab_switch'

    if violation_type == 'tab_switch':
        attempt.tab_switch_count += 1
        
        # Log detailed violation
        OnlineExamViolation.objects.create(
            attempt=attempt,
            event_type='tab_switch',
            details=_('Student switched tab/window. Total warnings: %d') % attempt.tab_switch_count
        )
        
        # Check if limit reached
        if attempt.cycle.enable_tab_warnings and attempt.tab_switch_count >= attempt.cycle.max_tab_switches:
            attempt.is_locked = True
            attempt.status = 'locked'
            attempt.lock_reason = _('Too many tab switches')
            from django.utils import timezone
            attempt.finished_at = timezone.now()
            attempt.save()
            return JsonResponse({
                'status': 'locked',
                'count': attempt.tab_switch_count,
            })
            
        attempt.save(update_fields=['tab_switch_count'])
        return JsonResponse({
            'status': 'success',
            'count': attempt.tab_switch_count,
            'limit': attempt.cycle.max_tab_switches if attempt.cycle.enable_tab_warnings else 0
        })

    return JsonResponse({'status': 'ok'})

@login_required
@user_passes_test(lambda u: is_student(u) or is_super_admin(u))
def cycle_exam_submit(request, attempt_pk):
    """Process student's submitted exam and calculate score."""
    if is_super_admin(request.user):
        attempt = get_object_or_404(OnlineAttempt, pk=attempt_pk)
    else:
        attempt = get_object_or_404(OnlineAttempt, pk=attempt_pk, student=request.user)
    
    if attempt.status == 'completed':
        return redirect('admissions:student_admission')
    
    # Accept locked attempts (auto-submitted due to violations)
    if attempt.status == 'locked' or attempt.is_locked:
        pass  # Allow processing to continue — will be set to 'completed' below
        
    tab_warning = request.POST.get('tab_warning')

    if request.method == 'POST' or attempt.status in ('timed_out', 'locked'):
        

        total_score = 0
        max_score = 0
        questions = attempt.cycle.admission_questions.prefetch_related('options')
        
        with transaction.atomic():
            for q in questions:
                max_score += q.points
                ans_id = request.POST.get(f'q_{q.id}')
                is_correct = False
                selected_option = None
                
                if ans_id:
                    try:
                        selected_option = q.options.get(id=ans_id)
                    except (AdmissionQuestionOption.DoesNotExist, ValueError):
                        pass
                
                # If still no selected_option, check if it was already saved (AJAX)
                if not selected_option:
                    saved_ans = attempt.answers.filter(question=q).first()
                    if saved_ans:
                        selected_option = saved_ans.selected_option
    
                if selected_option:
                    is_correct = selected_option.is_correct
                    if is_correct:
                        total_score += q.points
                        
                OnlineAttemptAnswer.objects.update_or_create(
                    attempt=attempt,
                    question=q,
                    defaults={
                        'selected_option': selected_option
                    }
                )
            
        percentage = (total_score / max_score * 100) if max_score > 0 else 0
        is_passed = percentage >= attempt.cycle.passing_score
        
        attempt.status = 'completed'
        from django.utils import timezone
        attempt.finished_at = timezone.now()
        # attempt only tracks status; final score goes to AdmissionResult

        attempt.save(update_fields=['status', 'finished_at'])
        
        # Save to AdmissionResult for analytics dashboard
        candidate = getattr(request.user, 'admission_profile', None)
        reg = AdmissionRegistration.objects.filter(user=request.user, cycle=attempt.cycle).first()
        if not reg:
            reg = AdmissionRegistration.objects.filter(user=request.user).order_by('-created_at').first()
        variant = reg.variant if reg else '1A'

        region = reg.region if reg else (candidate.region if candidate and hasattr(candidate, 'region') else '')
        region = region or ''
        
        school_name = ''
        if reg and getattr(reg, 'school_name', None):
            school_name = reg.school_name
        elif candidate and getattr(candidate, 'previous_school', None):
            school_name = candidate.previous_school.name

        # Find common phone numbers from registrations
        phone1 = reg.phone1 if reg else getattr(request.user, 'phone', '')
        phone2 = reg.phone2 if reg else getattr(request.user, 'father_phone', '')

        result, created = AdmissionResult.objects.update_or_create(
            exam_attempt=attempt,
            defaults={
                'cycle': attempt.cycle,
                'candidate': candidate,
                'variant': variant,
                'first_name': (reg.full_name.split(' ')[0] if reg and reg.full_name else request.user.first_name) or request.user.first_name,
                'last_name': (" ".join(reg.full_name.split(' ')[1:]) if reg and reg.full_name and len(reg.full_name.split(' ')) > 1 else request.user.last_name) or request.user.last_name,
                'admission_type': 'online',
                'score': float(total_score),
                'total_questions': int(max_score),
                'correct_count': int(total_score),
                'wrong_count': int(max_score - total_score),
                'percentage': float(percentage),
                'is_passed': is_passed,
                'region': str(region),
                'school_name': str(school_name),
                'phone1': phone1,
                'phone2': phone2,
                'tab_switch_count': attempt.tab_switch_count,
                'exam_duration_seconds': int((attempt.finished_at - attempt.started_at).total_seconds()) if attempt.finished_at else 0,
            }
        )
        
        # Subject Splits
        splits = attempt.cycle.subject_splits.filter(variant='1A')
        for split in splits:
            split_total = 0
            split_earned = 0
            split_qs = attempt.cycle.admission_questions.filter(
                order__gte=split.start_question,
                order__lte=split.end_question
            )
            for sq in split_qs:
                split_total += sq.points
                ans = attempt.answers.filter(question=sq).first()
                if ans and ans.selected_option and ans.selected_option.is_correct:
                    split_earned += sq.points
                    
            split_pct = (split_earned / split_total * 100) if split_total > 0 else 0
            AdmissionSubjectScore.objects.update_or_create(
                result=result,
                subject_name=split.subject_name,
                defaults={
                    'correct_count': int(split_earned),
                    'wrong_count': int(split_total - split_earned),
                    'total_questions': int(split_total),
                }
            )

        messages.success(request, _('Exam submitted successfully!'))
        return redirect('admissions:cycle_exam_result', attempt_pk=attempt.pk)
        
    return redirect('admissions:cycle_exam_take', attempt_pk=attempt.pk)

@login_required
@user_passes_test(lambda u: is_student(u) or is_super_admin(u))
def cycle_exam_result(request, attempt_pk):
    """View the results of the online admission exam."""
    # Determine back button URL
    if is_super_admin(request.user):
        attempt = get_object_or_404(OnlineAttempt, pk=attempt_pk)
        candidate = getattr(attempt.student, 'admission_profile', None)
        if candidate:
            back_url = reverse('admissions:candidate_detail', kwargs={'pk': candidate.pk})
        else:
            back_url = reverse('admissions:candidate_list')
    else:
        attempt = get_object_or_404(OnlineAttempt, pk=attempt_pk, student=request.user)
        back_url = reverse('admissions:student_admission')

    # Deny student detailed view if requested by user
    if not is_super_admin(request.user):
        messages.info(request, _('Your exam is completed. Results are being processed.'))
        return redirect('admissions:student_admission')

    answers = attempt.answers.select_related('question', 'selected_option').order_by('question__order')
    
    context = {
        'attempt': attempt,
        'cycle': attempt.cycle,
        'answers': answers,
        'back_url': back_url,
    }
    return render(request, 'admissions/cycle_exam_result.html', context)



# ============================================================
# MASTER ANSWER SETUP
# ============================================================

@login_required
@user_passes_test(is_super_admin)
def master_answer_setup(request, cycle_id=None):
    """Page to define/edit 4 master answer keys for a cycle."""
    if cycle_id:
        cycle = get_object_or_404(AdmissionCycle, pk=cycle_id)
    else:
        cycle = AdmissionCycle.objects.filter(is_active=True).first()
        if not cycle:
            cycle = AdmissionCycle.objects.first()
    
    if not cycle:
        messages.error(request, _("No admission cycle found."))
        return redirect('admissions:admission_analytics_dashboard')
    
    # Ensure subject splits exist
    _ensure_subject_splits(cycle)
    
    VARIANTS = ['1A', '1B', '2A', '2B']
    
    if request.method == 'POST':
        saved = 0
        for variant in VARIANTS:
            answer_key = request.POST.get(f'answers_{variant}', '')
            # Strip ALL whitespace (spaces, tabs, newlines) — user may paste from table
            answer_key = ''.join(answer_key.split())
            # Only keep first 70 chars
            answer_key = answer_key[:70]
            if answer_key:
                obj, created = AdmissionMasterAnswer.objects.update_or_create(
                    cycle=cycle,
                    variant=variant,
                    defaults={'answers': answer_key}
                )
                saved += 1

        
        messages.success(request, _('Saved %d master answer keys.') % saved)
        return redirect('admissions:master_answer_setup')
    
    # Load existing answers
    existing = {}
    for ma in AdmissionMasterAnswer.objects.filter(cycle=cycle):
        existing[ma.variant] = ma
    
    # Load subject splits for display
    splits = {}
    for split in AdmissionSubjectSplit.objects.filter(cycle=cycle):
        if split.variant not in splits:
            splits[split.variant] = []
        splits[split.variant].append(split)
    
    variant_data = []
    for v in VARIANTS:
        variant_data.append({
            'code': v,
            'answers': existing[v].answers if v in existing else '',
            'answer_count': existing[v].answer_count if v in existing else 0,
            'splits': splits.get(v, []),
        })
    
    context = {
        'cycle': cycle,
        'variant_data': variant_data,
        'cycles': AdmissionCycle.objects.all(),
    }
    return render(request, 'admissions/master_answer_setup.html', context)


# ============================================================
# RECALCULATE RESULTS
# ============================================================

@login_required
@user_passes_test(is_super_admin)
def recalculate_results(request, cycle_id=None):
    """Recalculate all results against current master answers."""
    if cycle_id:
        cycle = get_object_or_404(AdmissionCycle, pk=cycle_id)
    else:
        cycle = AdmissionCycle.objects.filter(is_active=True).first()
    
    if not cycle:
        messages.error(request, _("No cycle found."))
        return redirect('admissions:admission_analytics_dashboard')
    
    # Load master answers
    master_answers = {}
    for ma in AdmissionMasterAnswer.objects.filter(cycle=cycle):
        master_answers[ma.variant] = ma.answers[:70]
    
    if not master_answers:
        messages.error(request, _("No master answers defined. Please set them first."))
        return redirect('admissions:master_answer_setup')
    
    # Load universal subject splits (stored as 1A)
    universal_splits = list(AdmissionSubjectSplit.objects.filter(cycle=cycle, variant='1A'))
    
    results = AdmissionResult.objects.filter(cycle=cycle, answer_string__gt='')
    recalculated = 0
    
    for result in results:
        # Normalize variant for consistent lookup (Cyrillic → Latin)
        normalized_variant = _normalize_variant(result.variant)
        
        # Update the stored variant to normalized form if different
        if normalized_variant != result.variant:
            result.variant = normalized_variant
        
        if normalized_variant not in master_answers:
            continue
        
        master = master_answers[normalized_variant]
        answers_70 = result.answer_string[:70]
        
        correct_count = 0
        wrong_count = 0
        
        for i in range(min(70, len(answers_70), len(master))):
            student_char = answers_70[i]
            master_char = master[i]
            if student_char in ('-', '*', ' '):
                wrong_count += 1
            elif student_char == master_char:
                correct_count += 1
            else:
                wrong_count += 1
        
        if len(answers_70) < 70:
            wrong_count += 70 - len(answers_70)
        
        result.correct_count = correct_count
        result.wrong_count = wrong_count
        result.score = float(correct_count)
        result.percentage = (correct_count / 70 * 100)
        result.is_passed = result.percentage >= cycle.passing_score
        result.save()
        
        # Recalculate subject scores
        AdmissionSubjectScore.objects.filter(result=result).delete()
        if normalized_variant in master_answers:
            _calculate_subject_scores(result, answers_70, master, universal_splits)
        
        recalculated += 1
    
    messages.success(request, _('Recalculated %d results.') % recalculated)
    return redirect('admissions:admission_analytics_dashboard')


import time
from django.db import OperationalError
from functools import wraps

def retry_on_lock(max_retries=5, initial_wait=0.2):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except OperationalError as e:
                    if 'locked' in str(e).lower() and attempt < max_retries - 1:
                        time.sleep(initial_wait * (attempt + 1))
                        continue
                    raise
        return wrapper
    return decorator


@login_required
@retry_on_lock()
def save_answer_ajax(request, attempt_pk):
    """AJAX view to save a single answer during an active online attempt."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)
    if is_super_admin(request.user):
        attempt = get_object_or_404(OnlineAttempt, pk=attempt_pk, status__in=['in_progress', 'locked'])
    else:
        attempt = get_object_or_404(OnlineAttempt, pk=attempt_pk, student=request.user, status__in=['in_progress', 'locked'])
    
    question_id = request.POST.get('question_id')
    option_id = request.POST.get('option_id')
    
    if not question_id or not option_id:
        return JsonResponse({'status': 'error', 'message': 'Missing data'}, status=400)
    
    question = get_object_or_404(AdmissionQuestion, pk=question_id, cycle=attempt.cycle)
    option = get_object_or_404(AdmissionQuestionOption, pk=option_id, question=question)
    
    OnlineAttemptAnswer.objects.update_or_create(
        attempt=attempt,
        question=question,
        defaults={'selected_option': option}
    )
    
    return JsonResponse({'status': 'success'})





@login_required
@user_passes_test(lambda u: is_student(u) or is_super_admin(u))
def upload_recording_ajax(request, attempt_pk):
    """AJAX view to upload a webcam recording for an exam attempt."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)
    if is_super_admin(request.user):
        attempt = get_object_or_404(OnlineAttempt, pk=attempt_pk)
    else:
        attempt = get_object_or_404(OnlineAttempt, pk=attempt_pk, student=request.user)
    
    video_file = request.FILES.get('video')
    if not video_file:
        return JsonResponse({'status': 'error', 'message': 'No video file'}, status=400)
    
    duration = int(request.POST.get('duration', 0))
    
    recording = ExamRecording.objects.create(
        attempt=attempt,
        video_file=video_file,
        duration_seconds=duration,
        file_size_bytes=video_file.size,
    )
    
    return JsonResponse({'status': 'success', 'recording_id': recording.pk})


# ============================================================
# SUBJECT TOP STUDENTS
# ============================================================

@login_required
@user_passes_test(is_super_admin)
def subject_top_students(request):
    """Show top students filtered by subject."""
    cycle_id = request.GET.get('cycle')
    if cycle_id:
        active_cycle = get_object_or_404(AdmissionCycle, pk=cycle_id)
    else:
        active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
        if not active_cycle:
            active_cycle = AdmissionCycle.objects.first()
            
    all_cycles = AdmissionCycle.objects.all().order_by('-start_date')
    
    # Get available subjects
    all_subjects = AdmissionSubjectScore.objects.filter(
        result__cycle=active_cycle
    ).values_list('subject_name', flat=True).distinct().order_by('subject_name') if active_cycle else []
    
    selected_subject = request.GET.get('subject', '')
    
    scores = []
    if selected_subject and active_cycle:
        scores = AdmissionSubjectScore.objects.filter(
            result__cycle=active_cycle,
            subject_name=selected_subject
        ).select_related('result').order_by('-correct_count')[:100]
    
    context = {
        'all_cycles': all_cycles,
        'active_cycle': active_cycle,
        'subjects': list(all_subjects),
        'selected_subject': selected_subject,
        'scores': scores,
    }
    return render(request, 'admissions/subject_rankings.html', context)


# ============================================================
# SCHOOL SUBJECT ANALYTICS
# ============================================================

@login_required
@user_passes_test(is_super_admin)
def school_subject_analytics(request):
    """Show average correct/wrong per school for each subject."""
    cycle_id = request.GET.get('cycle')
    if cycle_id:
        active_cycle = get_object_or_404(AdmissionCycle, pk=cycle_id)
    else:
        active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
        if not active_cycle:
            active_cycle = AdmissionCycle.objects.first()
            
    all_cycles = AdmissionCycle.objects.all().order_by('-start_date')
    
    # Get all subjects
    all_subjects = list(
        AdmissionSubjectScore.objects.filter(
            result__cycle=active_cycle
        ).values_list('subject_name', flat=True).distinct().order_by('subject_name')
    ) if active_cycle else []
    
    selected_subject = request.GET.get('subject', '')
    
    school_data = []
    if selected_subject and active_cycle:
        from django.db.models import Avg
        school_data = AdmissionSubjectScore.objects.filter(
            result__cycle=active_cycle,
            subject_name=selected_subject
        ).values('result__school_name').annotate(
            avg_correct=Avg('correct_count'),
            avg_wrong=Avg('wrong_count'),
            student_count=Count('id'),
            total_q=models.Max('total_questions'),
        ).order_by('-avg_correct')
    
    context = {
        'all_cycles': all_cycles,
        'active_cycle': active_cycle,
        'subjects': all_subjects,
        'selected_subject': selected_subject,
        'school_data': school_data,
    }
    return render(request, 'admissions/school_subject_analytics.html', context)


# ============================================================
# ADMISSION REGISTRATION (Student self-registration)
# ============================================================

def admission_register(request):
    """Public registration form for admission candidates.
    
    Students enter their own school name (free text) and region.
    No user account required.
    The school is automatically added to the ExternalSchool registry.
    """
    active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
    
    if request.method == 'POST' and active_cycle:
        form = AdmissionRegistrationForm(request.POST)
        if form.is_valid():
            school_name = form.cleaned_data['school_name'].strip()
            
            registration = AdmissionRegistration.objects.create(
                cycle=active_cycle,
                full_name=form.cleaned_data['full_name'].strip(),
                gender=form.cleaned_data['gender'],
                school_name=school_name,
                region=form.cleaned_data['region'].strip(),
                phone1=form.cleaned_data['phone1'].strip(),
                phone2=form.cleaned_data.get('phone2', '').strip(),
                variant=form.cleaned_data['variant'],
                user=request.user if request.user.is_authenticated else None,
            )
            
            # Auto-create ExternalSchool if it doesn't exist
            _auto_create_external_school(school_name)
            
            return render(request, 'admissions/admission_register.html', {
                'success': True,
                'registration': registration,
                'active_cycle': active_cycle,
            })
    else:
        form = AdmissionRegistrationForm()
    
    return render(request, 'admissions/admission_register.html', {
        'form': form,
        'active_cycle': active_cycle,
    })


@login_required
@user_passes_test(is_super_admin)
def admission_registrations_list(request):
    """Admin view to see all admission registrations."""
    active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
    registrations = AdmissionRegistration.objects.filter(cycle=active_cycle) if active_cycle else AdmissionRegistration.objects.none()
    
    # Stats
    total = registrations.count()
    by_region = registrations.values('region').annotate(count=Count('id')).order_by('-count')
    by_variant = registrations.values('variant').annotate(count=Count('id')).order_by('variant')
    
    context = {
        'active_cycle': active_cycle,
        'registrations': registrations,
        'total': total,
        'by_region': by_region,
        'by_variant': by_variant,
    }
    return render(request, 'admissions/admission_registrations_list.html', context)
@login_required
@user_passes_test(is_super_admin)
def candidate_detail(request, pk):
    """Detailed view of an admission candidate."""
    candidate = get_object_or_404(AdmissionCandidate, pk=pk)
    
    # Get associated results and attempts
    results = AdmissionResult.objects.filter(candidate=candidate)
    online_attempts = OnlineAttempt.objects.filter(student=candidate.user)
    registrations = AdmissionRegistration.objects.filter(user=candidate.user)
    
    context = {
        'candidate': candidate,
        'results': results,
        'online_attempts': online_attempts,
        'registrations': registrations,
    }
    return render(request, 'admissions/candidate_detail.html', context)


@login_required
@user_passes_test(is_super_admin)
def online_analytics_dashboard(request):
    """Main dashboard for specifically online admission analytics."""
    active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
    if not active_cycle:
        active_cycle = AdmissionCycle.objects.first()
        
    if not active_cycle:
        messages.warning(request, _("No admission cycle found."))
        return redirect('admissions:dashboard')
        
    results = AdmissionResult.objects.filter(cycle=active_cycle, admission_type='online')
    total_count = results.count()
    
    avg_percentage = results.aggregate(avg=Avg('percentage'))['avg'] or 0
    passed_count = results.filter(is_passed=True).count()
    failed_count = results.filter(is_passed=False).count()
    
    medal_counts = {
        'gold': results.filter(percentage__gte=75).count(),
        'silver': results.filter(percentage__gte=60, percentage__lt=75).count(),
        'bronze': results.filter(percentage__gte=50, percentage__lt=60).count(),
        'none': results.filter(percentage__lt=50).count(),
    }
    
    context = {
        'cycle': active_cycle,
        'results': results,
        'total_count': total_count,
        'avg_percentage': avg_percentage,
        'passed_count': passed_count,
        'failed_count': failed_count,
        'medal_counts': medal_counts,
    }
    return render(request, 'admissions/online_analytics_dashboard.html', context)

@login_required
@user_passes_test(is_super_admin)
def online_subject_analytics(request):
    """Subject-based analysis for online admission."""
    active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
    if not active_cycle: active_cycle = AdmissionCycle.objects.first()
    
    splits = active_cycle.subject_splits.filter(variant='1A') if active_cycle else []
    subject_stats = []
    
    for split in splits:
        scores = AdmissionSubjectScore.objects.filter(
            result__cycle=active_cycle,
            result__admission_type='online',
            subject_name=split.subject_name
        )
        avg_pct = 0
        correct_sum = 0
        total_q_sum = 0
        if scores.exists():
            for s in scores:
                correct_sum += s.correct_count
                total_q_sum += s.total_questions
            avg_pct = (correct_sum / total_q_sum * 100) if total_q_sum > 0 else 0
            
        subject_stats.append({
            'name': split.subject_name,
            'count': scores.count(),
            'avg_percentage': avg_pct,
            'correct_sum': correct_sum,
            'total_qs': total_q_sum,
        })
        
    context = {
        'cycle': active_cycle,
        'subject_stats': subject_stats,
    }
    return render(request, 'admissions/online_subject_analytics.html', context)

@login_required
@user_passes_test(is_super_admin)
def online_school_analytics(request):
    """School-based analysis for online admission."""
    active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
    if not active_cycle: active_cycle = AdmissionCycle.objects.first()
    
    results = AdmissionResult.objects.filter(cycle=active_cycle, admission_type='online')
    school_names = results.values_list('school_name', flat=True).distinct()
    
    school_stats = []
    for school in school_names:
        school_results = results.filter(school_name=school)
        avg_pct = school_results.aggregate(avg=Avg('percentage'))['avg'] or 0
        total_in_school = school_results.count()
        pass_rate = (school_results.filter(is_passed=True).count() / total_in_school * 100) if total_in_school > 0 else 0
        
        school_stats.append({
            'name': school,
            'count': total_in_school,
            'avg_percentage': avg_pct,
            'pass_rate': pass_rate,
            'gold': school_results.filter(percentage__gte=75).count(),
            'silver': school_results.filter(percentage__gte=60, percentage__lt=75).count(),
            'bronze': school_results.filter(percentage__gte=50, percentage__lt=60).count(),
        })
        
    context = {
        'cycle': active_cycle,
        'school_stats': sorted(school_stats, key=lambda x: x['avg_percentage'], reverse=True),
    }
    return render(request, 'admissions/online_school_analytics.html', context)

@login_required
@user_passes_test(is_super_admin)
def online_top_students(request):
    """Top performers list for online admission."""
    active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
    if not active_cycle: active_cycle = AdmissionCycle.objects.first()
    
    results = AdmissionResult.objects.filter(
        cycle=active_cycle, 
        admission_type='online'
    ).order_by('-percentage', '-score')[:100]
    
    context = {
        'cycle': active_cycle,
        'results': results,
    }
    return render(request, 'admissions/online_top_students.html', context)

@login_required
@user_passes_test(is_super_admin)
def export_online_results_full(request):
    """Deep Excel export with registration + subject scores."""
    active_cycle = AdmissionCycle.objects.filter(is_active=True).first()
    if not active_cycle: active_cycle = AdmissionCycle.objects.first()
    
    results = AdmissionResult.objects.filter(cycle=active_cycle, admission_type='online').order_by('-percentage')
    subjects = list(active_cycle.subject_splits.values_list('subject_name', flat=True).distinct().order_by('subject_name'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Online Results"
    
    # Headers
    headers = [
        "Full Name", "Sex", "School", "Region", "Phone 1", "Phone 2", 
        "Total Questions", "Total Correct", "Percentage", "Passed", "Medal",
        "Tab Warnings", "Exam Status"
    ]
    
    # Find active survey questions across these attempts
    survey_qs = AdmissionQuestion.objects.filter(cycle=active_cycle, is_survey_question=True).order_by('order')
    survey_headers = [f"_{sq.question_text[:30]}" for sq in survey_qs]
    headers += survey_headers
    
    for s in subjects:
        headers += [f"{s} Correct", f"{s} Wrong", f"{s} Score %"]
        
    ws.append(headers)
    
    # Style
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
    for res in results:
        # Get registration info
        gender = ''
        phone1 = res.phone1
        phone2 = res.phone2
        reg = AdmissionRegistration.objects.filter(cycle=active_cycle, user=res.candidate.user).first() if (res.candidate and res.candidate.user) else None
        if reg:
            gender = reg.get_gender_display()
            phone1 = reg.phone1 or phone1
            phone2 = reg.phone2 or phone2
        elif res.candidate:
            gender = res.candidate.get_gender_display()
            
        row = [
            f"{res.first_name} {res.last_name}",
            gender,
            res.school_name,
            res.region,
            phone1,
            phone2,
            res.total_questions,
            res.correct_count,
            round(res.percentage, 1),
            "Yes" if res.is_passed else "No",
            res.medal.title() if res.medal else "",
            res.exam_attempt.tab_switch_count if hasattr(res, 'exam_attempt') and res.exam_attempt else 0,
            res.exam_attempt.get_status_display() if hasattr(res, 'exam_attempt') and res.exam_attempt else "N/A"
        ]
        
        # Add survey answers
        if hasattr(res, 'exam_attempt') and res.exam_attempt:
            attempt_answers = res.exam_attempt.answers.all()
            ans_map = {ans.question_id: ans.selected_option.text if ans.selected_option else "" for ans in attempt_answers}
            for sq in survey_qs:
                row.append(ans_map.get(sq.id, ""))
        else:
            for sq in survey_qs:
                row.append("")
        
        # Add subject details
        for s in subjects:
            ss = res.subject_scores.filter(subject_name=s).first()
            if ss:
                row += [ss.correct_count, ss.wrong_count, round(ss.percentage, 1)]
            else:
                row += ["", "", ""]
                
        ws.append(row)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=online_admission_full.xlsx'
    wb.save(response)
    return response

@login_required
@user_passes_test(is_super_admin)
def recalculate_all_online_results(request):
    """Utility to fix data issues in online admission results."""
    from .models import OnlineAttempt, AdmissionResult, AdmissionSubjectScore
    attempts = OnlineAttempt.objects.filter(status='completed')
    count = 0
    for attempt in attempts:
        # Re-run scoring logic
        total_score = 0
        max_score = 0
        questions = attempt.cycle.admission_questions.prefetch_related('options')
        
        # Track registration data
        candidate = getattr(attempt.student, 'admission_profile', None)
        reg = AdmissionRegistration.objects.filter(cycle=attempt.cycle, user=attempt.student).first()
        
        region = (reg.region if reg else '') or (candidate.region if candidate else '')
        phone1 = (reg.phone1 if reg else '') or (candidate.phone1 if candidate else '')
        phone2 = (reg.phone2 if reg else '') or (candidate.phone2 if candidate else '')
        school_name = (reg.school_name if reg else '') or (candidate.previous_school.name if candidate and candidate.previous_school else '')

        for q in questions:
            max_score += q.points
            ans = attempt.answers.filter(question=q).first()
            if ans and ans.selected_option and ans.selected_option.is_correct:
                total_score += q.points
        
        percentage = (total_score / max_score * 100) if max_score > 0 else 0
        is_passed = percentage >= attempt.cycle.passing_score
        
        result, created = AdmissionResult.objects.update_or_create(
            exam_attempt=attempt,
            defaults={
                'cycle': attempt.cycle,
                'candidate': candidate,
                'first_name': (reg.full_name.split(' ')[0] if reg and reg.full_name else attempt.student.first_name),
                'last_name': (" ".join(reg.full_name.split(' ')[1:]) if reg and reg.full_name and len(reg.full_name.split(' ')) > 1 else attempt.student.last_name),
                'admission_type': 'online',
                'score': float(total_score),
                'total_questions': int(max_score),
                'correct_count': int(total_score),
                'wrong_count': int(max_score - total_score),
                'percentage': float(percentage),
                'is_passed': is_passed,
                'region': region,
                'school_name': str(school_name),
                'phone1': phone1,
                'phone2': phone2,
            }
        )
        
        # Fix splits
        splits = attempt.cycle.subject_splits.filter(variant='1A')
        for split in splits:
            split_total = 0
            split_earned = 0
            split_qs = questions.filter(order__gte=split.start_question, order__lte=split.end_question)
            for sq in split_qs:
                split_total += sq.points
                ans = attempt.answers.filter(question=sq).first()
                if ans and ans.selected_option and ans.selected_option.is_correct:
                    split_earned += sq.points
            
            AdmissionSubjectScore.objects.update_or_create(
                result=result,
                subject_name=split.subject_name,
                defaults={
                    'correct_count': int(split_earned),
                    'wrong_count': int(split_total - split_earned),
                    'total_questions': int(split_total),
                }
            )
        count += 1
        
    messages.success(request, f"Successfully recalculated {count} student results.")
    return redirect('admissions:online_analytics_dashboard')


# ============================================================
# ROUND RESULTS — Admin management + Student-facing page
# ============================================================

@login_required
@user_passes_test(is_super_admin)
def round_results_admin(request):
    """Admin management page for round result sessions."""
    sessions = RoundResultSession.objects.all()
    context = {
        'sessions': sessions,
    }
    return render(request, 'admissions/round_results_admin.html', context)


@login_required
@user_passes_test(is_super_admin)
def round_results_upload(request):
    """Upload Excel and import round results."""
    if request.method == 'POST':
        title = request.POST.get('title', '1-тур жыйынтыгы')
        excel_file = request.FILES.get('file')
        try:
            passing_score = float(request.POST.get('passing_score', 24.0))
        except (ValueError, TypeError):
            passing_score = 24.0

        if not excel_file:
            messages.error(request, 'Excel файл тандалган жок.')
            return redirect('admissions:round_results_upload')

        session = RoundResultSession.objects.create(
            title=title,
            file=excel_file,
            uploaded_by=request.user,
            passing_score=passing_score,
        )

        import re

        def parse_subject(val):
            """Parse '8/10 (80.0%)' → (score=8, pct=80.0). Returns (None, None) on failure."""
            if not val:
                return None, None
            s = str(val).strip()
            # Try pattern: "8/10 (80.0%)"
            m = re.match(r'(\d+)\s*/\s*(\d+)\s*\((\d+\.?\d*)%\)', s)
            if m:
                score = float(m.group(1))
                pct = float(m.group(3))
                return score, pct
            # Fallback: just a number
            try:
                return float(s), None
            except (ValueError, TypeError):
                return None, None

        def safe_float(val):
            if val is None:
                return None
            try:
                s = str(val).replace('%', '').strip()
                return float(s)
            except (ValueError, TypeError):
                return None

        def safe_str(val):
            if val is None:
                return ''
            return str(val).strip()

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(row)
            wb.close()

            if len(rows) < 2:
                messages.warning(request, 'Excel файлда маалымат жок.')
                return redirect('admissions:round_results_admin')

            bulk_results = []
            count = 0
            for row in rows[1:]:
                if len(row) < 8:
                    continue
                name_val = row[0]
                if not name_val or str(name_val).strip() == '':
                    continue

                full_name = safe_str(name_val)
                gender = safe_str(row[1]) if len(row) > 1 else ''
                district = safe_str(row[2]) if len(row) > 2 else ''
                school = safe_str(row[3]) if len(row) > 3 else ''
                phone1 = safe_str(row[4]) if len(row) > 4 else ''
                phone2 = safe_str(row[5]) if len(row) > 5 else ''

                total_score_val = safe_float(row[6]) if len(row) > 6 else 0
                total_pct_val = safe_float(row[7]) if len(row) > 7 else 0

                # Parse subjects (columns 8-14)
                math_s, math_p = parse_subject(row[8]) if len(row) > 8 else (None, None)
                kyrgyz_s, kyrgyz_p = parse_subject(row[9]) if len(row) > 9 else (None, None)
                bio_s, bio_p = parse_subject(row[10]) if len(row) > 10 else (None, None)
                geo_s, geo_p = parse_subject(row[11]) if len(row) > 11 else (None, None)
                hist_s, hist_p = parse_subject(row[12]) if len(row) > 12 else (None, None)
                eng_s, eng_p = parse_subject(row[13]) if len(row) > 13 else (None, None)
                rus_s, rus_p = parse_subject(row[14]) if len(row) > 14 else (None, None)

                # Medal
                medal_raw = safe_str(row[15]) if len(row) > 15 else ''

                # Status logic: Score threshold OR medal
                status = 'rejected'
                passed_by_score = (total_score_val is not None and total_score_val >= passing_score)
                has_medal = False
                if medal_raw:
                    medal_lower = medal_raw.lower()
                    if any(k in medal_lower for k in ['gold', 'silver', 'bronze', 'алтын', 'күмүш', 'коло']):
                        has_medal = True
                
                if passed_by_score or has_medal:
                    status = 'accepted'

                bulk_results.append(RoundResult(
                    session=session,
                    full_name=full_name,
                    gender=gender,
                    district=district,
                    school=school,
                    phone1=phone1,
                    phone2=phone2,
                    total_score=total_score_val or 0,
                    total_pct=total_pct_val or 0,
                    math_score=math_s,
                    math_pct=math_p,
                    kyrgyz_score=kyrgyz_s,
                    kyrgyz_pct=kyrgyz_p,
                    biology_score=bio_s,
                    biology_pct=bio_p,
                    geography_score=geo_s,
                    geography_pct=geo_p,
                    history_score=hist_s,
                    history_pct=hist_p,
                    english_score=eng_s,
                    english_pct=eng_p,
                    russian_score=rus_s,
                    russian_pct=rus_p,
                    medal=medal_raw,
                    status=status,
                ))
                count += 1

            if bulk_results:
                RoundResult.objects.bulk_create(bulk_results, batch_size=500)

            messages.success(request, f'{count} окуучу ийгиликтүү жүктөлдү.')
        except Exception as e:
            messages.error(request, f'Excel файлды окууда ката: {e}')

        return redirect('admissions:round_results_admin')

    return render(request, 'admissions/round_results_upload.html')


@login_required
@user_passes_test(is_super_admin)
def round_results_toggle_publish(request, pk):
    """Toggle publish status of a round result session."""
    session = get_object_or_404(RoundResultSession, pk=pk)
    session.is_published = not session.is_published
    session.save()
    state = 'жарыяланды' if session.is_published else 'жашырылды'
    messages.success(request, f'"{session.title}" {state}.')
    return redirect('admissions:round_results_admin')


@login_required
@user_passes_test(is_super_admin)
def round_results_delete(request, pk):
    """Delete a round result session and all its results."""
    session = get_object_or_404(RoundResultSession, pk=pk)
    title = session.title
    session.delete()
    messages.success(request, f'"{title}" өчүрүлдү.')
    return redirect('admissions:round_results_admin')


@login_required
@user_passes_test(is_super_admin)
def round_results_session_edit(request, pk):
    """Page to edit names and schools of students in a session."""
    session = get_object_or_404(RoundResultSession, pk=pk)
    results = session.results.all().order_by('rank', 'full_name')
    context = {
        'session': session,
        'results': results,
    }
    return render(request, 'admissions/round_results_session_edit.html', context)


@login_required
@user_passes_test(is_super_admin)
@require_POST
def round_result_update_ajax(request, pk):
    """AJAX endpoint to update student name or school."""
    result = get_object_or_404(RoundResult, pk=pk)
    field = request.POST.get('field')
    value = request.POST.get('value', '').strip()

    if field == 'name':
        result.full_name = value
        result.save()
        return JsonResponse({'status': 'ok'})
    elif field == 'school':
        result.school = value
        result.save()
        return JsonResponse({'status': 'ok'})

    return JsonResponse({'status': 'error', 'message': 'Invalid field'}, status=400)


def round_results_student(request):
    """Student-facing page: top rankings + searchable results."""
    # Get latest published session
    session = RoundResultSession.objects.filter(is_published=True).first()
    if not session:
        return render(request, 'admissions/round_results_student.html', {
            'session': None,
        })

    all_results = session.results.all()

    # Top 10 per subject (exclude zero scores — 0/0 means subject wasn't taken)
    top_math = all_results.filter(math_score__isnull=False, math_score__gt=0).order_by('-math_score')[:10]
    top_kyrgyz = all_results.filter(kyrgyz_score__isnull=False, kyrgyz_score__gt=0).order_by('-kyrgyz_score')[:10]
    top_biology = all_results.filter(biology_score__isnull=False, biology_score__gt=0).order_by('-biology_score')[:10]
    top_geography = all_results.filter(geography_score__isnull=False, geography_score__gt=0).order_by('-geography_score')[:10]
    top_history = all_results.filter(history_score__isnull=False, history_score__gt=0).order_by('-history_score')[:10]
    top_english = all_results.filter(english_score__isnull=False, english_score__gt=0).order_by('-english_score')[:10]
    top_russian = all_results.filter(russian_score__isnull=False, russian_score__gt=0).order_by('-russian_score')[:10]
    top_total = all_results.order_by('-total_score')[:10]

    # Paginated + searchable results
    search_q = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    results_qs = all_results

    if search_q:
        results_qs = results_qs.filter(full_name__icontains=search_q)
    if status_filter:
        results_qs = results_qs.filter(status=status_filter)

    from django.core.paginator import Paginator
    paginator = Paginator(results_qs, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'session': session,
        'top_math': top_math,
        'top_kyrgyz': top_kyrgyz,
        'top_biology': top_biology,
        'top_geography': top_geography,
        'top_history': top_history,
        'top_english': top_english,
        'top_russian': top_russian,
        'top_total': top_total,
        'page_obj': page_obj,
        'search_q': search_q,
        'status_filter': status_filter,
        'total_count': all_results.count(),
        'accepted_count': all_results.filter(status='accepted').count(),
        'rejected_count': all_results.filter(status='rejected').count(),
    }
    return render(request, 'admissions/round_results_student.html', context)


def round_results_search_ajax(request):
    """AJAX endpoint for live search of round results."""
    from django.http import JsonResponse

    session = RoundResultSession.objects.filter(is_published=True).first()
    if not session:
        return JsonResponse({'results': [], 'total': 0, 'num_pages': 0,
                             'current_page': 1, 'has_prev': False, 'has_next': False})

    search_q = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    try:
        page = int(request.GET.get('page', 1))
    except (ValueError, TypeError):
        page = 1

    qs = session.results.all()
    if search_q:
        # SQLite icontains is case-insensitive for ASCII only.
        # For Cyrillic, annotate with Upper and filter.
        from django.db.models.functions import Upper
        qs = qs.annotate(_upper_name=Upper('full_name')).filter(
            _upper_name__icontains=search_q.upper()
        )
    if status_filter:
        qs = qs.filter(status=status_filter)

    from django.core.paginator import Paginator
    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(page)

    results = []
    start = page_obj.start_index()  # 1-based index
    for i, r in enumerate(page_obj):
        results.append({
            'id': r.id,
            'rank': start + i,
            'full_name': r.full_name or '',
            'school': r.school or '',
            'district': r.district or '',
            'total_score': float(r.total_score) if r.total_score else 0,
            'total_pct': float(r.total_pct) if r.total_pct else 0,
            'medal': r.medal if r.medal and r.medal.lower() != 'none' else '',
            'status': r.status or '',
        })

    return JsonResponse({
        'results': results,
        'total': paginator.count,
        'num_pages': paginator.num_pages,
        'current_page': page_obj.number,
        'has_prev': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
    })


def round_result_profile(request, pk):
    """Detailed profile page for a single student round result."""
    import json
    result = get_object_or_404(RoundResult, pk=pk)
    
    # We will pass the subjects as a JSON list to the template for Chart.js
    # Only include subjects that actually have a percentage or score recorded.
    subjects_data = []
    
    # helper for adding subjects if they exist
    def add_subject(label, score, pct):
        # Specific rule for language subjects: do not include if score is 0
        if label in ['Кыргыз тили', 'Орус тили']:
            s_val = float(score) if score is not None else 0.0
            p_val = float(pct) if pct is not None else 0.0
            if s_val == 0.0 and p_val == 0.0:
                return  # Skip adding this subject

        if score is not None or pct is not None:
            subjects_data.append({
                'label': label,
                'score': float(score) if score is not None else 0.0,
                'pct': float(pct) if pct is not None else 0.0
            })
            
    add_subject('Англис тили', result.english_score, result.english_pct)
    add_subject('Биология', result.biology_score, result.biology_pct)
    add_subject('Математика', result.math_score, result.math_pct)
    add_subject('Тарых', result.history_score, result.history_pct)
    add_subject('География', result.geography_score, result.geography_pct)
    add_subject('Кыргыз тили', result.kyrgyz_score, result.kyrgyz_pct)
    add_subject('Орус тили', result.russian_score, result.russian_pct)

    # Sort alphabet, or just leave as added (which is logical ordering often preferred)
    # We will pass this to the frontend.

    chart_labels = [s['label'] for s in subjects_data]
    chart_data = [s['pct'] for s in subjects_data]

    # Gender parsing for avatar
    g = (result.gender or '').strip().lower()
    is_boy = 'м' in g or 'эркек' in g
    is_girl = 'ж' in g or 'аял' in g or 'кыз' in g
    
    # If gender is not provided, guess from surname suffixes
    if not is_boy and not is_girl:
        full_name_upper = result.full_name.upper().strip()
        parts = full_name_upper.split()
        if parts:
            # Check the surname or patronymic (usually first or last part)
            last_name = parts[0]
            first_name = parts[-1] if len(parts) > 1 else ""
            patronymic = parts[1] if len(parts) > 2 else first_name
            
            # Kyrgyz and Russian male endings
            if last_name.endswith('ОВ') or last_name.endswith('ЕВ') or 'УУЛУ' in parts:
                is_boy = True
            # Kyrgyz and Russian female endings
            elif last_name.endswith('ОВА') or last_name.endswith('ЕВА') or 'КЫЗЫ' in parts:
                is_girl = True
            elif first_name.endswith('ОВ') or first_name.endswith('ЕВ') or patronymic.endswith('ОВ') or patronymic.endswith('ЕВ'):
                is_boy = True
            elif first_name.endswith('ОВА') or first_name.endswith('ЕВА') or patronymic.endswith('ОВА') or patronymic.endswith('ЕВА'):
                is_girl = True

    # Compute dynamic rank based on total_score
    all_students = result.session.results.all().order_by('-total_score', 'full_name')
    rank = None
    for index, student in enumerate(all_students, 1):
        if student.id == result.id:
            rank = index
            break

    context = {
        'result': result,
        'subjects_data': subjects_data,  # for the html cards/bars
        'chart_labels_json': json.dumps(chart_labels, ensure_ascii=False),
        'chart_data_json': json.dumps(chart_data),
        'is_boy': is_boy,
        'is_girl': is_girl,
        'rank': rank,
    }
    return render(request, 'admissions/round_result_profile.html', context)


@login_required
@user_passes_test(is_super_admin)
def exam_integrity_list(request):
    """Admin dashboard to list exam attempts with proctoring flags."""
    from django.core.paginator import Paginator
    from django.db.models import F, Sum
    
    cycle_id = request.GET.get('cycle_id')
    status_filter = request.GET.get('status')
    
    attempts = OnlineAttempt.objects.select_related('student', 'cycle').order_by('-started_at')
    
    if cycle_id:
        attempts = attempts.filter(cycle_id=cycle_id)
        
    if status_filter == 'locked':
        attempts = attempts.filter(status='locked')
        
    cycles = AdmissionCycle.objects.filter(is_active=True)
        
    paginator = Paginator(attempts, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'cycles': cycles,
        'selected_cycle': int(cycle_id) if cycle_id else '',
        'selected_status': status_filter or '',
    }
    return render(request, 'admissions/exam_integrity_list.html', context)


@login_required
@user_passes_test(is_super_admin)
def exam_integrity_detail(request, attempt_pk):
    """Detail view for a single attempt's integrity logs and recordings."""
    attempt = get_object_or_404(OnlineAttempt.objects.select_related('student', 'cycle'), pk=attempt_pk)
    
    recordings = attempt.recordings.all().order_by('uploaded_at')
    
    context = {
        'attempt': attempt,
        'recordings': recordings,
        'violations': attempt.violations.all(),
        'admission_result': getattr(attempt, 'admission_result', None),
    }
    return render(request, 'admissions/exam_integrity_detail.html', context)


@login_required
@user_passes_test(is_super_admin)
def cycle_recordings(request, pk):
    """Admin view to list all webcam recordings for a specific admission cycle."""
    from django.core.paginator import Paginator

    cycle = get_object_or_404(AdmissionCycle, pk=pk)
    recordings = ExamRecording.objects.filter(
        attempt__cycle=cycle
    ).select_related('attempt', 'attempt__student').order_by('-uploaded_at')

    # Delete recording
    if request.method == 'POST' and request.POST.get('delete_recording_id'):
        rec_id = request.POST.get('delete_recording_id')
        rec = get_object_or_404(ExamRecording, pk=rec_id, attempt__cycle=cycle)
        if rec.video_file:
            rec.video_file.delete(save=False)
        rec.delete()
        messages.success(request, _('Recording deleted successfully.'))
        return redirect('admissions:cycle_recordings', pk=pk)

    paginator = Paginator(recordings, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'cycle': cycle,
        'page_obj': page_obj,
        'total_recordings': recordings.count(),
    }
    return render(request, 'admissions/cycle_recordings.html', context)
