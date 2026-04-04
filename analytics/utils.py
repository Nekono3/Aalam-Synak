from django.db.models import Avg, Count, Max, Min, Q, Sum
from django.utils import timezone
from openpyxl.utils import get_column_letter
from exams.models import ExamAttempt, OnlineExam
from zipgrade.models import ExamResult

class AnalyticsHelper:
    """Helper class for analytics calculations."""

    @staticmethod
    def get_school_stats(school):
        """Get overall statistics for a school."""
        
        # Online Exam Stats
        online_attempts = ExamAttempt.objects.filter(
            exam__school=school, 
            status='completed'
        )
        
        online_stats = online_attempts.aggregate(
            total_attempts=Count('id'),
            avg_score=Avg('percentage'),
            max_score=Max('percentage'),
            min_score=Min('percentage')
        )
        
        # Calculate Pass Rate (assuming 60% is passing for now, or use exam specific)
        passed_attempts = 0
        if online_stats['total_attempts'] and online_stats['total_attempts'] > 0:
            # This is an approximation as passing_score is per exam
            # For a more accurate pass rate we need to check each attempt against its exam's passing score
            pass_count = 0
            for attempt in online_attempts.select_related('exam'):
                if attempt.percentage >= attempt.exam.passing_score:
                    pass_count += 1
            pass_rate = (pass_count / online_stats['total_attempts']) * 100
        else:
            pass_rate = 0

        # Subject Performance
        subject_performance = []
        # Group by subject name via Exam -> Subject
        # This is a bit complex in ORM across relationships, simplified approach:
        
        return {
            'online_exams': {
                'count': online_stats['total_attempts'] or 0,
                'avg_score': round(online_stats['avg_score'] or 0, 1),
                'pass_rate': round(pass_rate, 1),
                'max_score': round(online_stats['max_score'] or 0, 1),
                'min_score': round(online_stats['min_score'] or 0, 1),
            }
        }

    @staticmethod
    def get_online_exam_subject_performance(school):
        """Get subject performance stats for online exams."""
        from schools.models import Subject
        from django.db.models import Avg, Count
        
        subjects = Subject.objects.filter(online_exams__school=school).distinct()
        performance = []
        
        for subject in subjects:
            attempts = ExamAttempt.objects.filter(
                exam__school=school,
                exam__subject=subject,
                status='completed'
            )
            if attempts.exists():
                stats = attempts.aggregate(avg=Avg('percentage'), count=Count('id'))
                passed = sum(1 for a in attempts if a.percentage >= (a.exam.passing_score or 60))
                pass_rate = (passed / stats['count']) * 100
                
                performance.append({
                    'name': subject.name,
                    'count': stats['count'],
                    'avg_score': round(stats['avg'], 1),
                    'pass_rate': round(pass_rate, 1)
                })
        
        performance.sort(key=lambda x: x['avg_score'], reverse=True)
        return performance

    @staticmethod
    def get_all_schools_stats():
        """Get aggregated statistics across all schools."""
        
        # Online Exam Stats for all schools
        online_attempts = ExamAttempt.objects.filter(status='completed')
        
        online_stats = online_attempts.aggregate(
            total_attempts=Count('id'),
            avg_score=Avg('percentage'),
            max_score=Max('percentage'),
            min_score=Min('percentage')
        )
        
        # Calculate Pass Rate
        if online_stats['total_attempts'] and online_stats['total_attempts'] > 0:
            pass_count = 0
            for attempt in online_attempts.select_related('exam'):
                if attempt.percentage >= attempt.exam.passing_score:
                    pass_count += 1
            pass_rate = (pass_count / online_stats['total_attempts']) * 100
        else:
            pass_rate = 0

        return {
            'online_exams': {
                'count': online_stats['total_attempts'] or 0,
                'avg_score': round(online_stats['avg_score'] or 0, 1),
                'pass_rate': round(pass_rate, 1),
                'max_score': round(online_stats['max_score'] or 0, 1),
                'min_score': round(online_stats['min_score'] or 0, 1),
            }
        }

    @staticmethod
    def get_exam_stats(exam):
        """Get stats for a specific exam."""
        attempts = exam.attempts.filter(status='completed')
        
        total = attempts.count()
        if total == 0:
            return None
            
        stats = attempts.aggregate(
            avg=Avg('percentage'),
            max=Max('percentage'),
            min=Min('percentage')
        )
        
        passed = sum(1 for a in attempts if a.is_passed)
        
        return {
            'total_students': total,
            'avg_score': round(stats['avg'] or 0, 1),
            'max_score': round(stats['max'] or 0, 1),
            'min_score': round(stats['min'] or 0, 1),
            'pass_rate': round((passed / total) * 100, 1)
        }

    @staticmethod
    def get_classes_list(school):
        """Get distinct classes (grade + section) for a school."""
        from schools.models import MasterStudent
        
        classes = MasterStudent.objects.filter(school=school).values(
            'grade', 'section'
        ).distinct().order_by('grade', 'section')
        
        return [{'grade': c['grade'], 'section': c['section'], 
                 'name': f"{c['grade']}{c['section']}"} for c in classes]

    @staticmethod
    def get_class_stats(school, grade, section):
        """Get statistics for a specific class."""
        from schools.models import MasterStudent
        from accounts.models import User
        
        # Get all students in class from master list
        master_students = MasterStudent.objects.filter(
            school=school, grade=grade, section=section
        )
        
        # Get User accounts linked to this school that have taken exams
        # Match via similar full name patterns or student IDs
        student_ids = []
        for ms in master_students:
            # Try to find matching User by name pattern
            users = User.objects.filter(
                role='student',
                primary_school=school,
                first_name__icontains=ms.name,
                last_name__icontains=ms.surname
            )
            student_ids.extend([u.pk for u in users])
        
        if not student_ids:
            # Fallback: get all student users from the school
            student_ids = list(User.objects.filter(
                role='student',
                primary_school=school
            ).values_list('pk', flat=True))
        
        # Get exam attempts for these students
        attempts = ExamAttempt.objects.filter(
            student__pk__in=student_ids,
            status='completed'
        )
        
        total_attempts = attempts.count()
        if total_attempts == 0:
            return {
                'total_students': master_students.count(),
                'total_exams': 0,
                'avg_score': 0,
                'max_score': 0,
                'min_score': 0,
                'pass_rate': 0,
                'top_students': []
            }
        
        stats = attempts.aggregate(
            avg=Avg('percentage'),
            max=Max('percentage'),
            min=Min('percentage')
        )
        
        # Calculate pass rate
        passed_count = 0
        for attempt in attempts.select_related('exam'):
            if attempt.percentage >= attempt.exam.passing_score:
                passed_count += 1
        pass_rate = (passed_count / total_attempts) * 100 if total_attempts > 0 else 0
        
        # Get top performing students
        top_students = attempts.values(
            'student__first_name', 'student__last_name', 'student__pk'
        ).annotate(
            avg_score=Avg('percentage'),
            exams_taken=Count('id')
        ).order_by('-avg_score')[:5]
        
        return {
            'total_students': master_students.count(),
            'total_exams': total_attempts,
            'avg_score': round(stats['avg'] or 0, 1),
            'max_score': round(stats['max'] or 0, 1),
            'min_score': round(stats['min'] or 0, 1),
            'pass_rate': round(pass_rate, 1),
            'top_students': list(top_students)
        }

    @staticmethod
    def get_zipgrade_exams_for_school(school):
        """Get all ZipGrade exams for a school."""
        from zipgrade.models import ZipGradeExam
        return ZipGradeExam.objects.filter(school=school).order_by('-exam_date', '-created_at')

    @staticmethod
    def get_zipgrade_exam_stats(exam_ids):
        """Get aggregated statistics for selected ZipGrade exams."""
        from zipgrade.models import ZipGradeExam, ExamResult
        from decimal import Decimal
        
        if not exam_ids:
            return None
            
        results = ExamResult.objects.filter(exam_id__in=exam_ids)
        
        if not results.exists():
            return {
                'total_students': 0,
                'total_exams': len(exam_ids),
                'avg_score': 0,
                'max_score': 0,
                'min_score': 0,
                'pass_rate': 0,
                'exams_info': []
            }
        
        stats = results.aggregate(
            avg=Avg('percentage'),
            max=Max('percentage'),
            min=Min('percentage'),
            total=Count('id')
        )
        
        # Calculate pass rate (60% as passing threshold)
        passing_threshold = 60
        passed = results.filter(percentage__gte=passing_threshold).count()
        pass_rate = (passed / stats['total']) * 100 if stats['total'] > 0 else 0
        
        # Get exam info
        exams = ZipGradeExam.objects.filter(pk__in=exam_ids)
        exams_info = [{
            'id': e.pk,
            'title': e.title,
            'date': e.exam_date,
            'total_students': e.total_students,
            'avg_score': e.average_score
        } for e in exams]
        
        return {
            'total_students': stats['total'],
            'total_exams': len(exam_ids),
            'avg_score': round(float(stats['avg'] or 0), 1),
            'max_score': round(float(stats['max'] or 0), 1),
            'min_score': round(float(stats['min'] or 0), 1),
            'pass_rate': round(pass_rate, 1),
            'exams_info': exams_info
        }

    @staticmethod
    def get_zipgrade_class_breakdown(exam_ids, school):
        """Get breakdown by class for selected ZipGrade exams."""
        from zipgrade.models import ExamResult
        from schools.models import MasterStudent
        
        if not exam_ids:
            return []
        
        results = ExamResult.objects.filter(
            exam_id__in=exam_ids,
            student__isnull=False,
            student__school=school
        ).select_related('student')
        
        # Group by class (grade + section)
        class_stats = {}
        for result in results:
            if result.student:
                key = f"{result.student.grade}{result.student.section}"
                if key not in class_stats:
                    class_stats[key] = {
                        'grade': result.student.grade,
                        'section': result.student.section,
                        'name': key,
                        'scores': [],
                        'student_ids': set()
                    }
                class_stats[key]['scores'].append(float(result.percentage))
                class_stats[key]['student_ids'].add(result.student_id)
        
        # Calculate averages
        breakdown = []
        for key, data in class_stats.items():
            avg_score = sum(data['scores']) / len(data['scores']) if data['scores'] else 0
            passed = sum(1 for s in data['scores'] if s >= 60)
            pass_rate = (passed / len(data['scores'])) * 100 if data['scores'] else 0
            
            breakdown.append({
                'name': data['name'],
                'grade': data['grade'],
                'section': data['section'],
                'student_count': len(data['student_ids']),
                'avg_score': round(avg_score, 1),
                'pass_rate': round(pass_rate, 1),
                'max_score': round(max(data['scores']) if data['scores'] else 0, 1),
                'min_score': round(min(data['scores']) if data['scores'] else 0, 1)
            })
        
        # Sort by class name
        breakdown.sort(key=lambda x: (x['grade'], x['section']))
        return breakdown

    @staticmethod
    def get_zipgrade_subject_breakdown(exam_ids):
        """Get breakdown by subject for selected ZipGrade exams (if subject splits exist)."""
        from zipgrade.models import SubjectResult, SubjectSplit
        
        if not exam_ids:
            return []
        
        # Get all subject results for the selected exams
        subject_results = SubjectResult.objects.filter(
            result__exam_id__in=exam_ids
        ).select_related('subject_split__subject')
        
        if not subject_results.exists():
            return []
        
        # Group by subject
        subject_stats = {}
        for sr in subject_results:
            subject_name = sr.subject_split.subject.name
            if subject_name not in subject_stats:
                subject_stats[subject_name] = {
                    'name': subject_name,
                    'scores': []
                }
            subject_stats[subject_name]['scores'].append(float(sr.percentage))
        
        # Calculate aggregates
        breakdown = []
        for name, data in subject_stats.items():
            avg = sum(data['scores']) / len(data['scores']) if data['scores'] else 0
            passed = sum(1 for s in data['scores'] if s >= 60)
            pass_rate = (passed / len(data['scores'])) * 100 if data['scores'] else 0
            
            breakdown.append({
                'name': name,
                'student_count': len(data['scores']),
                'avg_score': round(avg, 1),
                'pass_rate': round(pass_rate, 1),
                'max_score': round(max(data['scores']) if data['scores'] else 0, 1),
                'min_score': round(min(data['scores']) if data['scores'] else 0, 1)
            })
        
        breakdown.sort(key=lambda x: x['name'])
        return breakdown

    @staticmethod
    def get_zipgrade_student_ranking(exam_ids, limit=20):
        """Get student ranking for selected ZipGrade exams."""
        from zipgrade.models import ExamResult
        
        if not exam_ids:
            return []
        
        results = ExamResult.objects.filter(
            exam_id__in=exam_ids
        ).select_related('student', 'exam')
        
        # Group by student
        student_stats = {}
        for result in results:
            # Use student ID if linked, otherwise ZipGrade ID
            if result.student:
                key = f"student_{result.student.pk}"
                name = result.student.full_name
                grade = result.student.grade
                section = result.student.section
            else:
                key = f"zg_{result.zipgrade_student_id}"
                name = result.display_name
                grade = '-'
                section = '-'
            
            if key not in student_stats:
                student_stats[key] = {
                    'name': name,
                    'grade': grade,
                    'section': section,
                    'scores': [],
                    'exams_taken': 0
                }
            student_stats[key]['scores'].append(float(result.percentage))
            student_stats[key]['exams_taken'] += 1
        
        # Calculate averages and sort
        ranking = []
        for key, data in student_stats.items():
            avg = sum(data['scores']) / len(data['scores']) if data['scores'] else 0
            ranking.append({
                'name': data['name'],
                'grade': data['grade'],
                'section': data['section'],
                'avg_score': round(avg, 1),
                'exams_taken': data['exams_taken'],
                'best_score': round(max(data['scores']) if data['scores'] else 0, 1),
                'worst_score': round(min(data['scores']) if data['scores'] else 0, 1)
            })
        
        # Sort by average score descending
        ranking.sort(key=lambda x: x['avg_score'], reverse=True)
        return ranking[:limit]

    @staticmethod
    def get_growth_chart_data(school=None):
        """Get growth chart data (last 12 weeks)."""
        from datetime import timedelta
        
        chart_labels = []
        chart_data = []
        today = timezone.now()
        
        filter_kwargs = {'status': 'completed'}
        if school:
            filter_kwargs['exam__school'] = school
            
        for i in range(11, -1, -1):
            start_date = today - timedelta(weeks=i+1)
            end_date = today - timedelta(weeks=i)
            
            # Copy dict to avoid reference issues
            current_filter = filter_kwargs.copy()
            current_filter['finished_at__range'] = (start_date, end_date)
            
            count = ExamAttempt.objects.filter(**current_filter).count()
            chart_labels.append(end_date.strftime('%d.%m'))
            chart_data.append(count)
            
        return chart_labels, chart_data

    @staticmethod
    def get_school_comparison_data(selected_exam_ids=None, selected_subject_id=None):
        """Get school comparison data (average scores)."""
        from schools.models import School
        from zipgrade.models import ZipGradeExam, SubjectResult
        
        labels = []
        data = []
        all_schools = School.objects.filter(is_active=True)
        
        for s in all_schools:
            s_exams = ZipGradeExam.objects.filter(school=s)
            if selected_exam_ids:
                s_exam_ids = [int(eid) for eid in selected_exam_ids if str(eid).isdigit()]
                s_exams = s_exams.filter(pk__in=s_exam_ids)
            
            s_exam_ids_list = list(s_exams.values_list('pk', flat=True))
            if s_exam_ids_list:
                if selected_subject_id:
                    subject_results = SubjectResult.objects.filter(
                        result__exam_id__in=s_exam_ids_list,
                        subject_split__subject_id=selected_subject_id
                    )
                    if subject_results.exists():
                        avg = sum(float(sr.percentage) for sr in subject_results) / subject_results.count()
                    else:
                        avg = 0
                else:
                    s_stats = AnalyticsHelper.get_zipgrade_exam_stats(s_exam_ids_list)
                    avg = s_stats['avg_score'] if s_stats else 0
            else:
                avg = 0
            labels.append(s.name)
            data.append(round(avg, 1))
            
        return labels, data

    @staticmethod
    def get_performance_distribution(exam_ids):
        """Get student performance distribution for PieChart.
        
        Categories:
        - Excellent: 80-100%
        - Good: 60-79%
        - Satisfactory: 50-59%
        - Unsatisfactory: 0-49%
        """
        from zipgrade.models import ExamResult
        from django.utils.translation import gettext as _
        
        if not exam_ids:
            return [], []
        
        results = ExamResult.objects.filter(exam_id__in=exam_ids)
        
        if not results.exists():
            return [], []
        
        # Count results in each category
        excellent = 0      # 80-100%
        good = 0           # 60-79%
        satisfactory = 0   # 50-59%
        unsatisfactory = 0 # 0-49%
        
        for result in results:
            percentage = float(result.percentage)
            if percentage >= 80:
                excellent += 1
            elif percentage >= 60:
                good += 1
            elif percentage >= 50:
                satisfactory += 1
            else:
                unsatisfactory += 1
        
        labels = [
            _('Excellent (80-100%)'),
            _('Good (60-79%)'),
            _('Satisfactory (50-59%)'),
            _('Unsatisfactory (0-49%)')
        ]
        data = [excellent, good, satisfactory, unsatisfactory]
        
        return labels, data

    @staticmethod
    def get_class_performance_distribution(exam_ids, grade, section):
        """Get performance distribution for a specific class in ZipGrade exams.
        
        Categories:
        - Excellent: 80-100%
        - Good: 60-79%
        - Satisfactory: 50-59%
        - Unsatisfactory: 0-49%
        """
        from zipgrade.models import ExamResult
        from django.utils.translation import gettext as _
        
        if not exam_ids:
            return [], []
        
        results = ExamResult.objects.filter(
            exam_id__in=exam_ids,
            student__grade=grade,
            student__section=section
        )
        
        if not results.exists():
            return [], []
        
        excellent = 0
        good = 0
        satisfactory = 0
        unsatisfactory = 0
        
        for result in results:
            percentage = float(result.percentage)
            if percentage >= 80:
                excellent += 1
            elif percentage >= 60:
                good += 1
            elif percentage >= 50:
                satisfactory += 1
            else:
                unsatisfactory += 1
        
        labels = [
            _('Excellent (80-100%)'),
            _('Good (60-79%)'),
            _('Satisfactory (50-59%)'),
            _('Unsatisfactory (0-49%)')
        ]
        data = [excellent, good, satisfactory, unsatisfactory]
        
        return labels, data

    @staticmethod
    def get_class_ranked_students(exam_ids, grade, section):
        """Get all students in a class ranked by average ZipGrade score (highest first).
        
        Returns list of dicts: [{name, avg_score, exam_count, percentage}, ...]
        """
        from zipgrade.models import ExamResult
        from django.db.models import Avg, Count
        
        if not exam_ids or not grade or not section:
            return []
        
        # Aggregate by student: average percentage and count of exams
        student_stats = (
            ExamResult.objects.filter(
                exam_id__in=exam_ids,
                student__grade=grade,
                student__section=section,
                student__isnull=False
            )
            .values('student__id', 'student__name', 'student__surname')
            .annotate(
                avg_score=Avg('percentage'),
                exam_count=Count('id')
            )
            .order_by('-avg_score')
        )
        
        ranked = []
        for i, s in enumerate(student_stats, 1):
            ranked.append({
                'rank': i,
                'name': f"{s['student__surname']} {s['student__name']}".strip(),
                'avg_score': round(float(s['avg_score']), 1),
                'exam_count': s['exam_count'],
            })
        
        return ranked

    @staticmethod
    def get_class_subject_breakdown(exam_ids, grade, section):
        """Get subject breakdown for a specific class in selected ZipGrade exams.
        
        Returns list of subjects sorted by average score (best performing first).
        Used for pie chart showing which subjects the class excels at.
        """
        from zipgrade.models import SubjectResult, ExamResult
        
        if not exam_ids or not grade or not section:
            return [], []
        
        # Get all exam results for students in this class
        class_results = ExamResult.objects.filter(
            exam_id__in=exam_ids,
            student__grade=grade,
            student__section=section
        ).values_list('id', flat=True)
        
        if not class_results:
            return [], []
        
        # Get subject results for these exam results
        subject_results = SubjectResult.objects.filter(
            result_id__in=class_results
        ).select_related('subject_split__subject')
        
        if not subject_results.exists():
            return [], []
        
        # Group by subject and calculate averages
        subject_stats = {}
        for sr in subject_results:
            subject_name = sr.subject_split.subject.name
            if subject_name not in subject_stats:
                subject_stats[subject_name] = {'scores': []}
            subject_stats[subject_name]['scores'].append(float(sr.percentage))
        
        # Calculate averages and build result
        breakdown = []
        for name, data in subject_stats.items():
            avg = sum(data['scores']) / len(data['scores']) if data['scores'] else 0
            breakdown.append({
                'name': name,
                'avg_score': round(avg, 1)
            })
        
        # Sort by average score descending (best subjects first)
        breakdown.sort(key=lambda x: x['avg_score'], reverse=True)
        
        # Return labels and data for pie chart
        labels = [s['name'] for s in breakdown]
        data = [s['avg_score'] for s in breakdown]
        
        return labels, data


class ReportGenerator:


    """Helper for generating analytic reports."""
    
    # Unicode font name constant — used in table styles and paragraphs
    UNICODE_FONT = 'LiberationSans'
    UNICODE_FONT_BOLD = 'LiberationSans-Bold'
    
    @staticmethod
    def _register_unicode_fonts():
        """Register Unicode-capable fonts for PDF generation.
        
        Returns (font_name, font_name_bold) tuple.
        Falls back to Helvetica if LiberationSans not available.
        """
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import os
        
        font_dir = '/usr/share/fonts/liberation'
        regular = os.path.join(font_dir, 'LiberationSans-Regular.ttf')
        bold = os.path.join(font_dir, 'LiberationSans-Bold.ttf')
        italic = os.path.join(font_dir, 'LiberationSans-Italic.ttf')
        
        if os.path.exists(regular):
            try:
                pdfmetrics.registerFont(TTFont('LiberationSans', regular))
                if os.path.exists(bold):
                    pdfmetrics.registerFont(TTFont('LiberationSans-Bold', bold))
                if os.path.exists(italic):
                    pdfmetrics.registerFont(TTFont('LiberationSans-Italic', italic))
                return 'LiberationSans', 'LiberationSans-Bold'
            except Exception:
                pass
        
        return 'Helvetica', 'Helvetica-Bold'
    
    @staticmethod
    def _get_unicode_styles():
        """Get stylesheet with Unicode font applied to all styles."""
        from reportlab.lib.styles import getSampleStyleSheet
        
        font_name, font_bold = ReportGenerator._register_unicode_fonts()
        styles = getSampleStyleSheet()
        
        # Override default fonts with Unicode-capable font
        for style_name in styles.byName:
            style = styles[style_name]
            if hasattr(style, 'fontName'):
                if 'Bold' in style.fontName or 'bold' in style.fontName.lower():
                    style.fontName = font_bold
                else:
                    style.fontName = font_name
        
        return styles, font_name, font_bold
    
    @staticmethod
    def generate_excel_report(school):
        """Generate Excel report for school analytics."""
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "School Analytics"
        
        # Header
        ws['A1'] = f"Analytics Report: {school.name}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Generated at:"
        ws['B3'] = timezone.now().strftime("%Y-%m-%d %H:%M")
        
        # Stats
        stats = AnalyticsHelper.get_school_stats(school)['online_exams']
        
        ws['A5'] = "Overview"
        ws['A5'].font = Font(bold=True)
        
        data = [
            ("Total Exams Taken", stats['count']),
            ("Average Score", f"{stats['avg_score']}%"),
            ("Pass Rate", f"{stats['pass_rate']}%"),
            ("Best Score", f"{stats['max_score']}%"),
        ]
        
        row = 6
        for label, value in data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
            
        # Subject Performance
        row += 2
        ws.cell(row=row, column=1, value="Subject Performance").font = Font(bold=True)
        row += 1
        headers = ["Subject", "Total Exams", "Avg Score", "Pass Rate"]
        for col, h in enumerate(headers, 1):
             ws.cell(row=row, column=col, value=h).font = Font(bold=True)
        row += 1
        
        subject_stats = AnalyticsHelper.get_online_exam_subject_performance(school)
        for subj in subject_stats:
            ws.cell(row=row, column=1, value=subj['name'])
            ws.cell(row=row, column=2, value=subj['count'])
            ws.cell(row=row, column=3, value=f"{subj['avg_score']}%")
            ws.cell(row=row, column=4, value=f"{subj['pass_rate']}%")
            row += 1
            
        # Recent Activity
        ws.cell(row=row+2, column=1, value="Recent Exams").font = Font(bold=True)
        row += 3
        
        headers = ["Title", "Subject", "Created By", "Date"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            
        exams = OnlineExam.objects.filter(school=school).order_by('-created_at')[:20]
        row += 1
        for exam in exams:
            ws.cell(row=row, column=1, value=exam.title)
            ws.cell(row=row, column=2, value=exam.subject.name)
            ws.cell(row=row, column=3, value=exam.created_by.get_full_name())
            ws.cell(row=row, column=4, value=exam.created_at.strftime("%Y-%m-%d"))
            row += 1
            
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=analytics_{school.pk}_{timezone.now().date()}.xlsx'
        wb.save(response)
        return response

    @staticmethod
    def generate_pdf_report(report_data):
        """Generate PDF report that matches the on-screen analytics data.
        
        report_data dict contains:
            - source: 'exams' or 'zipgrade'
            - school_name: str
            - stats: dict with count, avg_score, pass_rate, max_score
            For exams: growth_labels, growth_values
            For zipgrade: performance_distribution, class_comparison, recent_exams, subject_name
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from django.http import HttpResponse
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles, font_name, font_bold = ReportGenerator._get_unicode_styles()
        
        source = report_data.get('source', 'exams')
        school_name = report_data.get('school_name', 'Unknown')
        stats = report_data.get('stats', {})
        
        # Common table style
        header_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ])
        
        # Title
        source_label = 'ZipGrade' if source == 'zipgrade' else 'Exams'
        subject_name = report_data.get('subject_name', '')
        title_text = f"Analytics Report: {school_name} ({source_label})"
        if subject_name:
            title_text += f" — {subject_name}"
        
        elements.append(Paragraph(title_text, styles['Title']))
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # ── Overview Stats Table ──
        elements.append(Paragraph("Overview", styles['Heading2']))
        
        count_label = "Total Students" if source == 'zipgrade' else "Total Exams Taken"
        stats_data = [
            ["Metric", "Value"],
            [count_label, str(stats.get('count', 0))],
            ["Average Score", f"{stats.get('avg_score', 0)}%"],
            ["Pass Rate", f"{stats.get('pass_rate', 0)}%"],
            ["Best Score", f"{stats.get('max_score', 0)}%"],
        ]
        if stats.get('min_score') is not None and source == 'zipgrade':
            stats_data.append(["Lowest Score", f"{stats.get('min_score', 0)}%"])
        
        t = Table(stats_data, colWidths=[3*inch, 2*inch])
        t.setStyle(header_style)
        elements.append(t)
        elements.append(Spacer(1, 20))
        
        if source == 'zipgrade':
            # ── Performance Distribution ──
            perf_dist = report_data.get('performance_distribution', {})
            if perf_dist:
                elements.append(Paragraph("Performance Distribution", styles['Heading2']))
                perf_data = [["Category", "Students"]]
                for label, count in perf_dist.items():
                    perf_data.append([str(label), str(count)])
                
                pt = Table(perf_data, colWidths=[3*inch, 2*inch])
                pt.setStyle(header_style)
                elements.append(pt)
                elements.append(Spacer(1, 20))
            
            # ── Class Comparison ──
            class_comp = report_data.get('class_comparison', [])
            if class_comp:
                elements.append(Paragraph("Class Comparison (Average Score)", styles['Heading2']))
                class_data = [["Class", "Avg Score"]]
                for c in class_comp:
                    class_data.append([c.get('name', ''), f"{c.get('avg_score', 0)}%"])
                
                ct = Table(class_data, colWidths=[3*inch, 2*inch])
                ct.setStyle(header_style)
                elements.append(ct)
                elements.append(Spacer(1, 20))
            
            # ── Recent Exams ──
            recent = report_data.get('recent_exams', [])
            if recent:
                elements.append(Paragraph("Recent Exams", styles['Heading2']))
                exam_data = [["Title", "Date"]]
                for ex in recent:
                    date_str = ''
                    if ex.get('exam_date'):
                        date_str = ex['exam_date'].strftime('%Y-%m-%d') if hasattr(ex['exam_date'], 'strftime') else str(ex['exam_date'])
                    exam_data.append([str(ex.get('title', '')), date_str])
                
                et = Table(exam_data, colWidths=[4*inch, 1.5*inch])
                et.setStyle(header_style)
                elements.append(et)
                elements.append(Spacer(1, 20))
        
        else:
            # ── Exams: Growth Chart ──
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
            
            labels = report_data.get('growth_labels', [])
            values = report_data.get('growth_values', [])
            
            if values and any(v > 0 for v in values):
                elements.append(Paragraph("Exam Growth (Last 12 Weeks)", styles['Heading2']))
                d = Drawing(450, 200)
                lc = HorizontalLineChart()
                lc.x = 30
                lc.y = 50
                lc.height = 125
                lc.width = 400
                lc.data = [values]
                lc.categoryAxis.categoryNames = labels
                lc.categoryAxis.labels.boxAnchor = 'n'
                lc.categoryAxis.labels.angle = 30
                lc.categoryAxis.labels.dy = -10
                lc.valueAxis.valueMin = 0
                max_val = max(values) if values else 5
                lc.valueAxis.valueMax = max_val + (5 if max_val > 0 else 5)
                lc.valueAxis.valueStep = 5 if lc.valueAxis.valueMax > 15 else (1 if lc.valueAxis.valueMax < 5 else 2)
                d.add(lc)
                elements.append(d)
                elements.append(Spacer(1, 20))
        
        doc.build(elements)
        buffer.seek(0)
        
        # Use school name in filename
        safe_name = school_name.replace(' ', '_').replace('/', '_')[:30]
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=analytics_{safe_name}_{source}_{timezone.now().date()}.pdf'
        return response

    @staticmethod
    def generate_class_excel_report(school, grade, section):
        """Generate Excel report for class analytics."""
        import openpyxl
        from openpyxl.styles import Font
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Class {grade}{section} Analytics"
        
        # Header
        ws['A1'] = f"Class Analytics Report: {grade}{section}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"School: {school.name}"
        ws['A3'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Stats
        stats = AnalyticsHelper.get_class_stats(school, grade, section)
        
        ws['A5'] = "Overview"
        ws['A5'].font = Font(bold=True)
        
        data = [
            ("Total Students", stats['total_students']),
            ("Total Exams Taken", stats['total_exams']),
            ("Average Score", f"{stats['avg_score']}%"),
            ("Pass Rate", f"{stats['pass_rate']}%"),
            ("Best Score", f"{stats['max_score']}%"),
            ("Lowest Score", f"{stats['min_score']}%"),
        ]
        
        row = 6
        for label, value in data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Top students
        if stats['top_students']:
            row += 2
            ws.cell(row=row, column=1, value="Top Performers").font = Font(bold=True)
            row += 1
            headers = ["#", "Student Name", "Avg Score", "Exams"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h).font = Font(bold=True)
            row += 1
            for i, student in enumerate(stats['top_students'], 1):
                ws.cell(row=row, column=1, value=i)
                ws.cell(row=row, column=2, value=f"{student['student__first_name']} {student['student__last_name']}")
                ws.cell(row=row, column=3, value=f"{round(student['avg_score'], 1)}%")
                ws.cell(row=row, column=4, value=student['exams_taken'])
                row += 1
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=class_{grade}{section}_{timezone.now().date()}.xlsx'
        wb.save(response)
        return response

    @staticmethod
    def generate_class_pdf_report(report_data):
        """Generate PDF report for class analytics — matches on-screen data.
        
        report_data dict contains:
            - source: 'exams' or 'zipgrade'
            - school_name, class_name: str
            - stats: dict with total_students, total_exams, avg_score, pass_rate, max_score, min_score
            For zipgrade: performance_distribution, ranked_students
            For exams: stats dict from get_class_stats (includes top_students)
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from django.http import HttpResponse
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles, font_name, font_bold = ReportGenerator._get_unicode_styles()
        
        source = report_data.get('source', 'exams')
        school_name = report_data.get('school_name', '')
        class_name = report_data.get('class_name', '')
        stats = report_data.get('stats', {})
        
        # Common table style
        header_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), font_bold),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ])
        
        # Title
        source_label = 'ZipGrade' if source == 'zipgrade' else 'Exams'
        elements.append(Paragraph(f"Class Analytics: {class_name} ({source_label})", styles['Title']))
        elements.append(Paragraph(f"School: {school_name}", styles['Normal']))
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # ── Overview Stats ──
        elements.append(Paragraph("Overview", styles['Heading2']))
        stats_data = [
            ["Metric", "Value"],
            ["Students in Class", str(stats.get('total_students', 0))],
            ["Total Exams", str(stats.get('total_exams', 0))],
            ["Average Score", f"{stats.get('avg_score', 0)}%"],
            ["Pass Rate", f"{stats.get('pass_rate', 0)}%"],
            ["Best Score", f"{stats.get('max_score', 0)}%"],
            ["Lowest Score", f"{stats.get('min_score', 0)}%"],
        ]
        
        t = Table(stats_data, colWidths=[3*inch, 2*inch])
        t.setStyle(header_style)
        elements.append(t)
        elements.append(Spacer(1, 20))
        
        if source == 'zipgrade':
            # ── Performance Distribution ──
            perf_dist = report_data.get('performance_distribution', {})
            if perf_dist:
                elements.append(Paragraph("Performance Distribution", styles['Heading2']))
                perf_data = [["Category", "Students"]]
                for label, count in perf_dist.items():
                    perf_data.append([str(label), str(count)])
                pt = Table(perf_data, colWidths=[3*inch, 2*inch])
                pt.setStyle(header_style)
                elements.append(pt)
                elements.append(Spacer(1, 20))
            
            # ── Ranked Students ──
            ranked = report_data.get('ranked_students', [])
            if ranked:
                elements.append(Paragraph("Student Rankings (Highest to Lowest)", styles['Heading2']))
                rank_data = [["#", "Student Name", "Avg Score", "Exams"]]
                for s in ranked:
                    rank_data.append([
                        str(s.get('rank', '')),
                        s.get('name', ''),
                        f"{s.get('avg_score', 0)}%",
                        str(s.get('exam_count', 0))
                    ])
                rt = Table(rank_data, colWidths=[0.5*inch, 3*inch, 1*inch, 0.8*inch])
                rt.setStyle(header_style)
                # Highlight top 3
                for i in range(1, min(4, len(rank_data))):
                    rt.setStyle(TableStyle([
                        ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#d4edda')),
                    ]))
                elements.append(rt)
                elements.append(Spacer(1, 20))
        else:
            # ── Exams: Top Students ──
            top_students = stats.get('top_students', [])
            if top_students:
                elements.append(Paragraph("Top Performers", styles['Heading2']))
                top_data = [["#", "Student Name", "Avg Score", "Exams"]]
                for i, student in enumerate(top_students, 1):
                    name = f"{student.get('student__name', '')} {student.get('student__surname', '')}".strip()
                    top_data.append([
                        str(i),
                        name,
                        f"{round(student.get('avg_score', 0), 1)}%",
                        str(student.get('exams_taken', 0))
                    ])
                tt = Table(top_data, colWidths=[0.5*inch, 3*inch, 1*inch, 0.8*inch])
                tt.setStyle(header_style)
                elements.append(tt)
                elements.append(Spacer(1, 20))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=class_{class_name}_{source}_{timezone.now().date()}.pdf'
        return response

    @staticmethod
    def generate_student_excel_report(student):
        """Generate Excel report for student analytics."""
        import openpyxl
        from openpyxl.styles import Font
        from django.http import HttpResponse
        from django.db.models import Avg, Max, Min
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Analytics"
        
        # Header
        ws['A1'] = f"Student Analytics Report: {student.get_full_name()}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Email: {student.email}"
        ws['A3'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Get attempts
        attempts = ExamAttempt.objects.filter(student=student, status='completed').order_by('-started_at')
        
        if attempts.exists():
            stats = attempts.aggregate(
                avg=Avg('percentage'),
                max=Max('percentage'),
                min=Min('percentage')
            )
            passed = sum(1 for a in attempts if a.percentage >= (a.exam.passing_score or 60))
            
            ws['A5'] = "Overview"
            ws['A5'].font = Font(bold=True)
            
            data = [
                ("Total Exams", attempts.count()),
                ("Passed Exams", passed),
                ("Average Score", f"{round(stats['avg'] or 0, 1)}%"),
                ("Best Score", f"{round(stats['max'] or 0, 1)}%"),
                ("Lowest Score", f"{round(stats['min'] or 0, 1)}%"),
            ]
            
            row = 6
            for label, value in data:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value)
                row += 1
            
            # Exam history
            row += 2
            ws.cell(row=row, column=1, value="Exam History").font = Font(bold=True)
            row += 1
            headers = ["Exam", "Subject", "Score", "Date", "Status"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=h).font = Font(bold=True)
            row += 1
            for attempt in attempts[:20]:
                ws.cell(row=row, column=1, value=attempt.exam.title)
                ws.cell(row=row, column=2, value=attempt.exam.subject.name)
                ws.cell(row=row, column=3, value=f"{attempt.percentage}%")
                ws.cell(row=row, column=4, value=attempt.started_at.strftime("%Y-%m-%d"))
                ws.cell(row=row, column=5, value="Passed" if attempt.is_passed else "Failed")
                row += 1
        else:
            ws['A5'] = "No exam data available"
        
        # Auto-adjust columns
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=student_{student.pk}_{timezone.now().date()}.xlsx'
        wb.save(response)
        return response

    @staticmethod
    def generate_student_pdf_report(student):
        """Generate PDF report for student analytics."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from django.http import HttpResponse
        from django.db.models import Avg, Max, Min
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles, font_name, font_bold = ReportGenerator._get_unicode_styles()
        
        # Title
        elements.append(Paragraph(f"Student Analytics Report: {student.get_full_name()}", styles['Title']))
        elements.append(Paragraph(f"Email: {student.email}", styles['Normal']))
        elements.append(Paragraph(f"Date: {timezone.now().strftime('%Y-%m-%d')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Get attempts
        attempts = ExamAttempt.objects.filter(student=student, status='completed').order_by('-started_at')
        
        if attempts.exists():
            # Get data for chart (reverse chronologically)
            chart_labels = [a.exam.title[:15] for a in reversed(attempts[:15])]
            chart_values = [float(a.percentage) for a in reversed(attempts[:15])]
            
            stats = attempts.aggregate(
                avg=Avg('percentage'),
                max=Max('percentage'),
                min=Min('percentage')
            )
            passed = sum(1 for a in attempts if a.percentage >= (a.exam.passing_score or 60))
            
            data = [
                ["Metric", "Value"],
                ["Total Exams", str(attempts.count())],
                ["Passed Exams", str(passed)],
                ["Average Score", f"{round(stats['avg'] or 0, 1)}%"],
                ["Best Score", f"{round(stats['max'] or 0, 1)}%"],
                ["Lowest Score", f"{round(stats['min'] or 0, 1)}%"],
            ]
            
            t = Table(data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_bold),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 20))
            
            # Exam History Chart
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
            
            elements.append(Paragraph("Exam Performance History", styles['Heading2']))
            
            if chart_values:
                d = Drawing(400, 200)
                lc = HorizontalLineChart()
                lc.x = 30
                lc.y = 50
                lc.height = 125
                lc.width = 350
                lc.data = [chart_values]
                lc.categoryAxis.categoryNames = chart_labels
                lc.categoryAxis.labels.boxAnchor = 'n'
                lc.categoryAxis.labels.angle = 30
                lc.categoryAxis.labels.dy = -10
                lc.valueAxis.valueMin = 0
                lc.valueAxis.valueMax = 100
                lc.valueAxis.valueStep = 20
                d.add(lc)
                elements.append(d)
                elements.append(Spacer(1, 20))
            
            # Recent exams
            elements.append(Paragraph("Recent Exam History", styles['Heading2']))
            exam_data = [["Exam", "Score", "Date", "Status"]]
            for attempt in attempts[:10]:
                exam_data.append([
                    attempt.exam.title[:30],
                    f"{attempt.percentage}%",
                    attempt.started_at.strftime("%Y-%m-%d"),
                    "Passed" if attempt.is_passed else "Failed"
                ])
            t2 = Table(exam_data)
            t2.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_bold),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(t2)
        else:
            elements.append(Paragraph("No exam data available for this student.", styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=student_{student.pk}_{timezone.now().date()}.pdf'
        return response
